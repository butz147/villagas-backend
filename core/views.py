from datetime import datetime, timedelta, date
from decimal import Decimal, InvalidOperation
import json

import requests
from openpyxl import Workbook
from openpyxl.styles import Font

from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.core.paginator import Paginator
from django.db import models
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.views.decorators.csrf import csrf_exempt

from core.models import Produto
from core.models import Loja
from .models import (
    AuditoriaSistema,
    Cliente,
    CompraEstoque,
    ContaPagar,
    ContaReceber,
    Comodato,
    Despesa,
    Entregador,
    FechamentoCaixa,
    InventarioEstoque,
    ItemInventarioEstoque,
    Loja,
    MovimentacaoEstoque,
    Pedido,
    PerfilUsuario,
    Produto,
    RetiradaFuncionario,
    Venda,
    ValeGas,
    Veiculo,
    AbastecimentoVeiculo,
    ManutencaoVeiculo,
    VendaAntecipada,
    CupomDesconto,
    Fornecedor,
    VeiculoRota,
    RotaEntrega,
    MetricaEntregador,
    ChecklistVeiculo,
    AlertaManutencao,
    Comissao,
    Meta_Vendas,
    NotificacaoSistema,
)


def obter_loja_ativa(request):
    """
    Retorna a loja ativa do usuário considerando seleção multi-loja via sessão.
    Admins com múltiplos perfis podem trocar de loja sem sair do sistema.
    """
    loja_id = request.session.get('loja_ativa_id')
    if loja_id:
        try:
            # Confirma que o usuário tem acesso a esta loja
            perfil = PerfilUsuario.objects.get(user=request.user, loja_id=loja_id)
            return perfil.loja
        except PerfilUsuario.DoesNotExist:
            pass
    # Fallback: primeira loja do perfil
    perfil = PerfilUsuario.objects.filter(user=request.user).first()
    if perfil and perfil.loja:
        request.session['loja_ativa_id'] = perfil.loja.id
        return perfil.loja
    return None


def obter_loja_usuario(user):
    perfil = PerfilUsuario.objects.filter(user=user).first()
    return perfil.loja if perfil else None


def obter_perfil_usuario(user):
    return PerfilUsuario.objects.filter(user=user).first()


def usuario_eh_funcionario(user):
    perfil = obter_perfil_usuario(user)
    return perfil and perfil.tipo_usuario == "funcionario"


def usuario_eh_gerente(user):
    perfil = obter_perfil_usuario(user)
    return perfil and perfil.tipo_usuario == "gerente"


def usuario_eh_admin(user):
    perfil = obter_perfil_usuario(user)
    return user.is_staff or (perfil and perfil.tipo_usuario == "admin")


def usuario_eh_gerente_ou_admin(user):
    perfil = obter_perfil_usuario(user)
    return user.is_staff or (perfil and perfil.tipo_usuario in ["gerente", "admin"])


def dia_fechado(loja, data=None):
    if data is None:
        data = timezone.now().date()

    return FechamentoCaixa.objects.filter(loja=loja, data=data).exists()


def somar_pagamentos_mistos(vendas):
    dinheiro = 0
    pix = 0
    credito = 0
    debito = 0

    for venda in vendas:
        if venda.forma_pagamento_1 == "dinheiro":
            dinheiro += float(venda.valor_pagamento_1 or 0)
        elif venda.forma_pagamento_1 == "pix":
            pix += float(venda.valor_pagamento_1 or 0)
        elif venda.forma_pagamento_1 == "credito":
            credito += float(venda.valor_pagamento_1 or 0)
        elif venda.forma_pagamento_1 == "debito":
            debito += float(venda.valor_pagamento_1 or 0)

        if venda.forma_pagamento_2 == "dinheiro":
            dinheiro += float(venda.valor_pagamento_2 or 0)
        elif venda.forma_pagamento_2 == "pix":
            pix += float(venda.valor_pagamento_2 or 0)
        elif venda.forma_pagamento_2 == "credito":
            credito += float(venda.valor_pagamento_2 or 0)
        elif venda.forma_pagamento_2 == "debito":
            debito += float(venda.valor_pagamento_2 or 0)

    return dinheiro, pix, credito, debito


def somar_pagamentos_pedidos(pedidos):
    dinheiro = 0
    pix = 0
    credito = 0
    debito = 0

    for pedido in pedidos:
        valor = float(pedido.quantidade * pedido.preco_unitario)

        if pedido.forma_pagamento == "dinheiro":
            dinheiro += valor
        elif pedido.forma_pagamento == "pix":
            pix += valor
        elif pedido.forma_pagamento == "credito":
            credito += valor
        elif pedido.forma_pagamento == "debito":
            debito += valor

    return dinheiro, pix, credito, debito

def enviar_para_google_sheets(venda):
    url = "COLE_AQUI_SEU_LINK_DO_GOOGLE_SCRIPT"

    pagamento_texto = f"{venda.forma_pagamento_1} - R$ {venda.valor_pagamento_1}"
    if venda.forma_pagamento_2 and venda.valor_pagamento_2:
        pagamento_texto += f" | {venda.forma_pagamento_2} - R$ {venda.valor_pagamento_2}"

    dados = {
        "data": venda.data_venda.strftime("%d/%m/%Y"),
        "hora": venda.data_venda.strftime("%H:%M"),
        "loja": venda.loja.nome if venda.loja else "",
        "cliente": venda.cliente.nome if venda.cliente else "Sem cliente",
        "produto": venda.produto.nome,
        "tipo_venda": venda.get_tipo_venda_display(),
        "quantidade": venda.quantidade,
        "preco": str(venda.preco_unitario),
        "pagamento": pagamento_texto,
        "funcionario": venda.funcionario.username
    }

    try:
        requests.post(url, json=dados, timeout=10)
    except Exception as e:
        print("Erro ao enviar para Google Sheets:", e)


def lista_produtos(request):
    return render(request, 'produtos.html')


def registrar_auditoria(loja, usuario, acao, descricao=""):
    AuditoriaSistema.objects.create(
        loja=loja,
        usuario=usuario,
        acao=acao,
        descricao=descricao,
    )

def login_usuario(request):
    if request.user.is_authenticated:
        if usuario_eh_admin(request.user):
            return redirect("/admin-geral/")
        elif usuario_eh_gerente(request.user):
            return redirect("/")
        else:
            return redirect("/")

    erro = ""

    if request.method == "POST":
        username = request.POST.get("username")
        password = request.POST.get("password")
        eh_folguista = request.POST.get("eh_folguista") == "1"
        folguista_nome = request.POST.get("folguista_nome", "").strip()

        user = authenticate(request, username=username, password=password)

        if user is not None:
            login(request, user)

            if eh_folguista and folguista_nome:
                request.session["folguista_ativo"] = True
                request.session["folguista_nome"] = folguista_nome
            else:
                request.session["folguista_ativo"] = False
                request.session["folguista_nome"] = ""

            if usuario_eh_admin(user):
                return redirect("/admin-geral/")
            elif usuario_eh_gerente(user):
                return redirect("/")
            else:
                return redirect("/")
        else:
            erro = "Usuário ou senha inválidos."

    return render(request, "login.html", {"erro": erro})


@login_required
def logout_usuario(request):
    logout(request)
    return redirect("/login/")


@login_required
def inicio(request):
    loja = obter_loja_usuario(request.user)

    if not loja:
        return render(request, "erro_loja.html")

    if usuario_eh_admin(request.user):
        return redirect("/admin-geral/")

    if usuario_eh_gerente(request.user):
        return render(request, "home_gerente.html", {
            "loja": loja,
        })

    return render(request, "home_funcionario.html", {
        "loja": loja,
    })

@login_required
def registrar_venda(request):
    loja = obter_loja_usuario(request.user)

    if not loja:
        return render(request, "erro_loja.html")

    if dia_fechado(loja):
        return render(request, "operacao_bloqueada.html", {
            "mensagem": "O dia de hoje já foi fechado. Não é possível registrar novas vendas."
        })

    produtos = Produto.objects.filter(loja=loja)
    clientes = Cliente.objects.filter(loja=loja).order_by("nome")

    if request.method == "POST":
        produto_id = request.POST.get("produto")
        cliente_id = request.POST.get("cliente")
        quantidade = int(request.POST.get("quantidade"))
        preco = request.POST.get("preco")

        forma_pagamento_1 = request.POST.get("forma_pagamento_1")
        valor_pagamento_1 = request.POST.get("valor_pagamento_1")

        forma_pagamento_2 = request.POST.get("forma_pagamento_2")
        valor_pagamento_2 = request.POST.get("valor_pagamento_2")

        tipo_venda = request.POST.get("tipo_venda", "normal")

        produto = Produto.objects.get(id=produto_id, loja=loja)

        # Gás do Povo: valor definido pelo preco_venda do produto no admin
        if forma_pagamento_1 == "gas_do_povo":
            valor_pagamento_1 = str(produto.preco_venda)
            valor_pagamento_2 = None
            forma_pagamento_2 = None
        elif forma_pagamento_2 == "gas_do_povo":
            valor_pagamento_2 = str(produto.preco_venda)

        cliente = None
        if cliente_id:
            cliente = Cliente.objects.get(id=cliente_id, loja=loja)

        try:
            preco_decimal = Decimal(preco)
            total_venda = Decimal(quantidade) * preco_decimal

            valor_1 = Decimal(valor_pagamento_1) if valor_pagamento_1 else Decimal("0.00")
            valor_2 = Decimal(valor_pagamento_2) if valor_pagamento_2 else Decimal("0.00")
        except (InvalidOperation, TypeError):
            return render(request, "venda.html", {
                "produtos": produtos,
                "clientes": clientes,
                "loja": loja,
                "erro": "Valores inválidos."
            })

        if not forma_pagamento_2:
            forma_pagamento_2 = None
            valor_2 = Decimal("0.00")

        soma_pagamentos = (valor_1 + valor_2).quantize(Decimal("0.01"))
        total_venda = total_venda.quantize(Decimal("0.01"))

        if soma_pagamentos != total_venda:
            return render(request, "venda.html", {
                "produtos": produtos,
                "clientes": clientes,
                "loja": loja,
                "erro": f"A soma dos pagamentos deve ser igual ao total da venda (R$ {total_venda})."
            })

        if valor_1 <= 0:
            return render(request, "venda.html", {
                "produtos": produtos,
                "clientes": clientes,
                "loja": loja,
                "erro": "O valor do pagamento 1 deve ser maior que zero."
            })

        if forma_pagamento_2 and valor_2 <= 0:
            return render(request, "venda.html", {
                "produtos": produtos,
                "clientes": clientes,
                "loja": loja,
                "erro": "O valor do pagamento 2 deve ser maior que zero."
            })

        if not produto.controla_retorno:
            tipo_venda = "normal"

            if produto.estoque_cheio < quantidade:
                return render(request, "venda.html", {
                    "produtos": produtos,
                    "clientes": clientes,
                    "loja": loja,
                    "erro": "Estoque insuficiente."
                })
        else:
            if tipo_venda == "troca":
                if produto.estoque_cheio < quantidade:
                    return render(request, "venda.html", {
                        "produtos": produtos,
                        "clientes": clientes,
                        "loja": loja,
                        "erro": "Estoque cheio insuficiente para troca."
                    })

            elif tipo_venda == "completo":
                if produto.estoque_cheio < quantidade:
                    return render(request, "venda.html", {
                        "produtos": produtos,
                        "clientes": clientes,
                        "loja": loja,
                        "erro": "Estoque cheio insuficiente para venda completa."
                    })

            elif tipo_venda == "casco":
                if produto.estoque_vazio < quantidade:
                    return render(request, "venda.html", {
                        "produtos": produtos,
                        "clientes": clientes,
                        "loja": loja,
                        "erro": "Estoque vazio insuficiente para venda de casco."
                    })

            else:
                return render(request, "venda.html", {
                    "produtos": produtos,
                    "clientes": clientes,
                    "loja": loja,
                    "erro": "Tipo de venda inválido para produto retornável."
                })

        venda = Venda.objects.create(
            funcionario=request.user,
            loja=loja,
            cliente=cliente,
            produto=produto,
            quantidade=quantidade,
            preco_unitario=preco_decimal,
            forma_pagamento_1=forma_pagamento_1,
            valor_pagamento_1=valor_1,
            forma_pagamento_2=forma_pagamento_2,
            valor_pagamento_2=valor_2 if forma_pagamento_2 else None,
            tipo_venda=tipo_venda,
        )

        if not produto.controla_retorno:
            produto.estoque_cheio -= quantidade

            MovimentacaoEstoque.objects.create(
                loja=loja,
                produto=produto,
                usuario=request.user,
                tipo="saida",
                quantidade=quantidade,
                motivo="Venda normal de produto simples"
            )
        else:
            if tipo_venda == "troca":
                produto.estoque_cheio -= quantidade
                produto.estoque_vazio += quantidade

                MovimentacaoEstoque.objects.create(
                    loja=loja,
                    produto=produto,
                    usuario=request.user,
                    tipo="saida",
                    quantidade=quantidade,
                    motivo="Venda por troca - saída de cheio"
                )

            elif tipo_venda == "completo":
                produto.estoque_cheio -= quantidade

                MovimentacaoEstoque.objects.create(
                    loja=loja,
                    produto=produto,
                    usuario=request.user,
                    tipo="saida",
                    quantidade=quantidade,
                    motivo="Venda completa - saída de cheio sem retorno de vazio"
                )

            elif tipo_venda == "casco":
                produto.estoque_vazio -= quantidade

                MovimentacaoEstoque.objects.create(
                    loja=loja,
                    produto=produto,
                    usuario=request.user,
                    tipo="saida",
                    quantidade=quantidade,
                    motivo="Venda de casco - saída de vazio"
                )

        produto.save()

        enviar_para_google_sheets(venda)

        registrar_auditoria(
            loja=loja,
            usuario=request.user,
            acao="Venda registrada",
            descricao=f"Venda #{venda.id} | Produto: {produto.nome} | Quantidade: {quantidade} | Tipo: {tipo_venda}"
        )

        return redirect("/venda/")

    return render(request, "venda.html", {
        "produtos": produtos,
        "clientes": clientes,
        "loja": loja
    })

@login_required
def dashboard(request):
    if not usuario_eh_gerente_ou_admin(request.user):
        return render(request, "operacao_bloqueada.html", {
            "mensagem": "Apenas gerente ou admin podem acessar o dashboard."
        })

    loja = obter_loja_usuario(request.user)

    if not loja:
        return render(request, "erro_loja.html")

    hoje = timezone.now().date()
    filtro = request.GET.get("filtro", "7dias")

    data_inicial = hoje - timedelta(days=6)
    data_final = hoje

    if filtro == "hoje":
        data_inicial = hoje
        data_final = hoje
    elif filtro == "ontem":
        data_inicial = hoje - timedelta(days=1)
        data_final = data_inicial
    elif filtro == "7dias":
        data_inicial = hoje - timedelta(days=6)
        data_final = hoje
    elif filtro == "30dias":
        data_inicial = hoje - timedelta(days=29)
        data_final = hoje
    elif filtro == "mes":
        data_inicial = hoje.replace(day=1)
        data_final = hoje
    elif filtro == "personalizado":
        data_ini_str = request.GET.get("data_inicial")
        data_fim_str = request.GET.get("data_final")

        if data_ini_str and data_fim_str:
            data_inicial = datetime.strptime(data_ini_str, "%Y-%m-%d").date()
            data_final = datetime.strptime(data_fim_str, "%Y-%m-%d").date()

    vendas_periodo = Venda.objects.filter(
        data_venda__date__gte=data_inicial,
        data_venda__date__lte=data_final,
        loja=loja,
        status="ativa"
    )

    total_periodo = 0
    for venda in vendas_periodo:
        total_periodo += float(venda.quantidade * venda.preco_unitario)

    dinheiro, pix, credito, debito = somar_pagamentos_mistos(vendas_periodo)

    total_vendas = vendas_periodo.count()

    produtos_vendidos = {}
    for venda in vendas_periodo:
        nome = venda.produto.nome
        produtos_vendidos[nome] = produtos_vendidos.get(nome, 0) + venda.quantidade

    produto_mais_vendido = "Nenhum"
    qtd_produto_mais_vendido = 0

    if produtos_vendidos:
        produto_mais_vendido = max(produtos_vendidos, key=produtos_vendidos.get)
        qtd_produto_mais_vendido = produtos_vendidos[produto_mais_vendido]

    produtos = Produto.objects.filter(loja=loja)

    estoque_baixo = []
    for produto in produtos:
        if produto.estoque_cheio < produto.alerta_estoque_minimo:
            estoque_baixo.append(produto)

    resumo_funcionarios = {}

    for venda in vendas_periodo:
        nome_funcionario = venda.funcionario.username

        if nome_funcionario not in resumo_funcionarios:
            resumo_funcionarios[nome_funcionario] = {
                "quantidade_vendas": 0,
                "valor_total": 0
            }

        resumo_funcionarios[nome_funcionario]["quantidade_vendas"] += 1
        resumo_funcionarios[nome_funcionario]["valor_total"] += float(
            venda.quantidade * venda.preco_unitario
        )

    ranking_funcionarios = sorted(
        resumo_funcionarios.items(),
        key=lambda item: item[1]["valor_total"],
        reverse=True
    )

    labels = []
    valores = []

    dia_atual = data_inicial
    while dia_atual <= data_final:
        total_dia = 0
        vendas_dia = vendas_periodo.filter(data_venda__date=dia_atual)
        for venda in vendas_dia:
            total_dia += float(venda.quantidade * venda.preco_unitario)

        labels.append(dia_atual.strftime("%d/%m"))
        valores.append(float(total_dia))
        dia_atual += timedelta(days=1)

    return render(request, "dashboard.html", {
        "loja": loja,
        "total": total_periodo,
        "dinheiro": dinheiro,
        "pix": pix,
        "credito": credito,
        "debito": debito,
        "total_vendas_hoje": total_vendas,
        "produto_mais_vendido": produto_mais_vendido,
        "qtd_produto_mais_vendido": qtd_produto_mais_vendido,
        "produtos": produtos,
        "estoque_baixo": estoque_baixo,
        "ranking_funcionarios": ranking_funcionarios,
        "labels_grafico": labels,
        "valores_grafico": valores,
        "filtro": filtro,
        "data_inicial": data_inicial,
        "data_final": data_final,
    })

@login_required
def relatorio_diario(request):
    loja = obter_loja_usuario(request.user)

    if not loja:
        return render(request, "erro_loja.html")

    hoje = timezone.now().date()
    filtro = request.GET.get("filtro", "hoje")

    data_inicial = hoje
    data_final = hoje

    if filtro == "ontem":
        data_inicial = hoje - timedelta(days=1)
        data_final = data_inicial

    elif filtro == "7dias":
        data_inicial = hoje - timedelta(days=6)
        data_final = hoje

    elif filtro == "mes":
        data_inicial = hoje.replace(day=1)
        data_final = hoje

    elif filtro == "personalizado":
        data_ini_str = request.GET.get("data_inicial")
        data_fim_str = request.GET.get("data_final")

        if data_ini_str and data_fim_str:
            data_inicial = datetime.strptime(data_ini_str, "%Y-%m-%d").date()
            data_final = datetime.strptime(data_fim_str, "%Y-%m-%d").date()

    if request.user.is_staff:
        vendas = Venda.objects.filter(
            data_venda__date__gte=data_inicial,
            data_venda__date__lte=data_final,
            loja=loja
        ).order_by("-data_venda")
    else:
        vendas = Venda.objects.filter(
            data_venda__date__gte=data_inicial,
            data_venda__date__lte=data_final,
            loja=loja,
            funcionario=request.user
        ).order_by("-data_venda")

    total = 0
    for venda in vendas:
        total += venda.quantidade * venda.preco_unitario

    contexto = {
        "vendas": vendas,
        "total_geral": total,
        "data_inicial": data_inicial,
        "data_final": data_final,
        "loja": loja,
        "filtro": filtro,
    }

    return render(request, "relatorio.html", contexto)


@login_required
def movimentar_estoque(request):
    loja = obter_loja_usuario(request.user)

    if not loja:
        return render(request, "erro_loja.html")

    if not usuario_eh_gerente_ou_admin(request.user):
        return render(request, "operacao_bloqueada.html", {
            "mensagem": "Apenas gerente ou admin podem movimentar estoque."
        })

    if dia_fechado(loja):
        return render(request, "operacao_bloqueada.html", {
            "mensagem": "O dia de hoje já foi fechado. Não é possível movimentar estoque."
        })

    produtos = Produto.objects.filter(loja=loja).order_by("nome")
    movimentacoes = MovimentacaoEstoque.objects.filter(loja=loja).order_by("-data_movimentacao")[:50]
    erro = ""

    if request.method == "POST":
        produto_id = request.POST.get("produto")
        tipo = request.POST.get("tipo")
        quantidade = request.POST.get("quantidade")
        motivo = request.POST.get("motivo", "").strip()

        try:
            quantidade = int(quantidade)
        except (TypeError, ValueError):
            erro = "Quantidade inválida."
            return render(request, "movimentacao_estoque.html", {
                "loja": loja,
                "produtos": produtos,
                "movimentacoes": movimentacoes,
                "erro": erro,
            })

        if quantidade <= 0:
            erro = "A quantidade deve ser maior que zero."
            return render(request, "movimentacao_estoque.html", {
                "loja": loja,
                "produtos": produtos,
                "movimentacoes": movimentacoes,
                "erro": erro,
            })

        try:
            produto = Produto.objects.get(id=produto_id, loja=loja)
        except Produto.DoesNotExist:
            erro = "Produto não encontrado."
            return render(request, "movimentacao_estoque.html", {
                "loja": loja,
                "produtos": produtos,
                "movimentacoes": movimentacoes,
                "erro": erro,
            })

        if tipo == "ajuste_cheio":
            produto.estoque_cheio = quantidade

        elif tipo == "ajuste_vazio":
            produto.estoque_vazio = quantidade

        else:
            erro = "Tipo de correção inválido."
            return render(request, "movimentacao_estoque.html", {
                "loja": loja,
                "produtos": produtos,
                "movimentacoes": movimentacoes,
                "erro": erro,
            })

        produto.save()

        MovimentacaoEstoque.objects.create(
            loja=loja,
            produto=produto,
            usuario=request.user,
            tipo=tipo,
            quantidade=quantidade,
            motivo=motivo,
        )
        registrar_auditoria(
            loja=loja,
            usuario=request.user,
            acao="Correção de estoque",
            descricao=f"Produto: {produto.nome} | Tipo: {tipo} | Quantidade: {quantidade} | Motivo: {motivo}"
        )
        
        return redirect("/movimentacao-estoque/")

    return render(request, "movimentacao_estoque.html", {
        "loja": loja,
        "produtos": produtos,
        "movimentacoes": movimentacoes,
        "erro": erro,
    })

@login_required
def pedidos(request):
    loja = obter_loja_usuario(request.user)

    if not loja:
        return render(request, "erro_loja.html")

    produtos = Produto.objects.filter(loja=loja).order_by("nome")
    clientes = Cliente.objects.filter(loja=loja, ativo=True).order_by("nome")
    entregadores = Entregador.objects.filter(loja=loja).order_by("nome")

    status_choices = [
        ("novo", "Novo"),
        ("preparando", "Preparando"),
        ("saiu_entrega", "Saiu para entrega"),
        ("entregue", "Entregue"),
        ("cancelado", "Cancelado"),
    ]

    if request.method == "POST":
        if dia_fechado(loja):
            return render(request, "operacao_bloqueada.html", {
                "mensagem": "O dia de hoje já foi fechado. Não é possível criar ou alterar pedidos."
            })

        try:
            produto_id = request.POST.get("produto")
            cliente_id = request.POST.get("cliente")
            cliente_nome_novo = request.POST.get("cliente_nome_novo", "").strip()
            cliente_telefone_novo = request.POST.get("cliente_telefone_novo", "").strip()
            cliente_endereco_novo = request.POST.get("cliente_endereco_novo", "").strip()
            quantidade = int(request.POST.get("quantidade") or 1)
            preco_unitario = request.POST.get("preco_unitario", "").strip()
            forma_pagamento = request.POST.get("forma_pagamento")
            frete = request.POST.get("frete", "0").strip() or "0"
            observacoes = request.POST.get("observacoes", "").strip()
            entregador_id = request.POST.get("entregador", "").strip()

            produto = Produto.objects.filter(id=produto_id, loja=loja).first()
            if not produto:
                return render(request, "pedidos.html", {
                    "loja": loja,
                    "produtos": produtos,
                    "clientes": clientes,
                    "entregadores": entregadores,
                    "status_choices": status_choices,
                    "erro": "Produto não encontrado."
                })

            # Gás do Povo: preço vem do produto, não do formulário
            if forma_pagamento == "gas_do_povo":
                preco_unitario = str(produto.preco_venda)

            if not preco_unitario:
                return render(request, "pedidos.html", {
                    "loja": loja,
                    "produtos": produtos,
                    "clientes": clientes,
                    "entregadores": entregadores,
                    "status_choices": status_choices,
                    "erro": "Informe o preço unitário."
                })

            try:
                preco_unitario = Decimal(preco_unitario.replace(",", "."))
            except:
                return render(request, "pedidos.html", {
                    "loja": loja,
                    "produtos": produtos,
                    "clientes": clientes,
                    "entregadores": entregadores,
                    "status_choices": status_choices,
                    "erro": "Preço unitário inválido."
                })

            if quantidade <= 0:
                return render(request, "pedidos.html", {
                    "loja": loja,
                    "produtos": produtos,
                    "clientes": clientes,
                    "entregadores": entregadores,
                    "status_choices": status_choices,
                    "erro": "Quantidade inválida."
                })

            cliente = None

            if cliente_id:
                try:
                    cliente = Cliente.objects.get(id=cliente_id, loja=loja)
                except Cliente.DoesNotExist:
                    cliente = None

            elif cliente_nome_novo:
                cliente_existente = None

                if cliente_telefone_novo:
                    cliente_existente = Cliente.objects.filter(
                        loja=loja,
                        telefone=cliente_telefone_novo
                    ).first()

                if cliente_existente:
                    cliente = cliente_existente
                else:
                    cliente = Cliente.objects.create(
                        loja=loja,
                        nome=cliente_nome_novo,
                        telefone=cliente_telefone_novo,
                        endereco=cliente_endereco_novo,
                        ativo=True,
                    )

            entregador = None
            if entregador_id:
                try:
                    entregador = Entregador.objects.get(id=entregador_id, loja=loja)
                except Entregador.DoesNotExist:
                    entregador = None

            pedido = Pedido.objects.create(
                loja=loja,
                cliente=cliente,
                produto=produto,
                entregador=entregador,
                quantidade=quantidade,
                preco_unitario=preco_unitario,
                frete=Decimal(frete),
                forma_pagamento=forma_pagamento,
                status="novo",
                observacoes=observacoes,
            )

            registrar_auditoria(
                loja=loja,
                usuario=request.user,
                acao="Pedido criado",
                descricao=f"Pedido #{pedido.id} criado para {cliente.nome if cliente else 'sem cliente'}"
            )

            messages.success(request, f"Pedido #{pedido.id} criado com sucesso.")
            return redirect("pedidos")

        except Exception as e:
            return render(request, "pedidos.html", {
                "loja": loja,
                "produtos": produtos,
                "clientes": clientes,
                "entregadores": entregadores,
                "status_choices": status_choices,
                "erro": f"Erro ao criar pedido: {str(e)}"
            })

    busca = request.GET.get("q", "").strip()
    status_filtro = request.GET.get("status", "").strip()
    pagina_atual = request.GET.get("page", 1)

    pedidos_qs = Pedido.objects.filter(loja=loja).select_related(
        "cliente", "produto", "entregador"
    )

    if busca:
        pedidos_qs = pedidos_qs.filter(
            models.Q(cliente__nome__icontains=busca) |
            models.Q(cliente__telefone__icontains=busca) |
            models.Q(produto__nome__icontains=busca) |
            models.Q(observacoes__icontains=busca)
        )

    if status_filtro:
        pedidos_qs = pedidos_qs.filter(status=status_filtro)

    pedidos_qs = pedidos_qs.order_by("-data_pedido")

    total_pedidos = pedidos_qs.count()

    paginator = Paginator(pedidos_qs, 15)
    pedidos_page = paginator.get_page(pagina_atual)

    return render(request, "pedidos.html", {
        "loja": loja,
        "produtos": produtos,
        "clientes": clientes,
        "entregadores": entregadores,
        "pedidos": pedidos_page,
        "status_choices": status_choices,
        "status_filtro": status_filtro,
        "busca": busca,
        "total_pedidos": total_pedidos,
    })


@login_required
def buscar_clientes_ajax(request):
    loja = obter_loja_usuario(request.user)

    if not loja:
        return JsonResponse({"resultados": []})

    termo = request.GET.get("q", "").strip()

    clientes = Cliente.objects.filter(loja=loja, ativo=True)

    if termo:
        clientes = clientes.filter(
            models.Q(nome__icontains=termo) |
            models.Q(telefone__icontains=termo)
        )

    clientes = clientes.order_by("nome")[:10]

    resultados = [
        {
            "id": cliente.id,
            "nome": cliente.nome,
            "telefone": cliente.telefone or "",
            "endereco": cliente.endereco or "",
        }
        for cliente in clientes
    ]

    return JsonResponse({"resultados": resultados})


def obter_loja_por_cidade(cidade):
    cidade = (cidade or "").strip().lower()
    return Loja.objects.filter(cidade__iexact=cidade).first()


def loja_esta_aberta(loja):
    if not loja or not loja.horario_abertura or not loja.horario_fechamento:
        return True

    agora = timezone.localtime().time()
    abertura = loja.horario_abertura
    fechamento = loja.horario_fechamento

    if abertura <= fechamento:
        return abertura <= agora <= fechamento

    return agora >= abertura or agora <= fechamento


def api_ok(dados=None, mensagem=""):
    return JsonResponse({
        "ok": True,
        "mensagem": mensagem,
        "dados": dados or {},
    })


def api_error(erro, status=400):
    return JsonResponse({
        "ok": False,
        "erro": erro,
    }, status=status)


def listar_produtos_site(request):
    cidade = request.GET.get("cidade")

    if isinstance(cidade, list):
        cidade = cidade[0] if cidade else None

    if not cidade or not isinstance(cidade, str):
        return api_error("Cidade inválida", status=400)

    cidade = cidade.strip().lower()

    loja = Loja.objects.filter(cidade__iexact=cidade).first()

    if not loja:
        return api_error("Loja não encontrada", status=404)

    produtos = Produto.objects.filter(loja=loja).order_by("nome")

    dados = {
        "produtos": [
            {
                "id": p.id,
                "nome": p.nome,
                "preco_venda": str(p.preco_venda),
                "imagem_url": p.imagem.url if p.imagem else ""
            }
            for p in produtos
        ],
        "loja_aberta": loja_esta_aberta(loja),
        "horario_abertura": loja.horario_abertura.strftime("%H:%M") if loja.horario_abertura else "",
        "horario_fechamento": loja.horario_fechamento.strftime("%H:%M") if loja.horario_fechamento else "",
        "nome_loja": loja.nome,
        "cidade": loja.cidade,
        "tempo_entrega_min": loja.tempo_entrega_min,
        "tempo_entrega_max": loja.tempo_entrega_max,
    }

    return JsonResponse({
        "ok": True,
        "mensagem": "Produtos carregados com sucesso.",
        "dados": dados,
        # compatibilidade temporária com frontend atual
        **dados,
    })


def validar_cupom(codigo, subtotal):
    codigo = (codigo or "").strip().upper()

    if not codigo:
        return {
            "valido": True,
            "cupom": None,
            "desconto": Decimal("0.00"),
            "erro": "",
        }

    cupom = CupomDesconto.objects.filter(codigo__iexact=codigo, ativo=True).first()

    if not cupom:
        return {
            "valido": False,
            "cupom": None,
            "desconto": Decimal("0.00"),
            "erro": "Cupom inválido.",
        }

    hoje = date.today()

    if cupom.data_inicio and hoje < cupom.data_inicio:
        return {
            "valido": False,
            "cupom": None,
            "desconto": Decimal("0.00"),
            "erro": "Cupom ainda não está disponível.",
        }

    if cupom.data_fim and hoje > cupom.data_fim:
        return {
            "valido": False,
            "cupom": None,
            "desconto": Decimal("0.00"),
            "erro": "Cupom expirado.",
        }

    if cupom.uso_maximo is not None and cupom.total_usado >= cupom.uso_maximo:
        return {
            "valido": False,
            "cupom": None,
            "desconto": Decimal("0.00"),
            "erro": "Cupom esgotado.",
        }

    subtotal_decimal = Decimal(str(subtotal or 0))

    if cupom.tipo_desconto == "percentual":
        desconto = (subtotal_decimal * Decimal(str(cupom.valor_desconto))) / Decimal("100")
    else:
        desconto = Decimal(str(cupom.valor_desconto))

    if desconto > subtotal_decimal:
        desconto = subtotal_decimal

    return {
        "valido": True,
        "cupom": cupom,
        "desconto": desconto.quantize(Decimal("0.01")),
        "erro": "",
    }


@csrf_exempt
def criar_pedido_site(request):
    if request.method != "POST":
        return api_error("Método não permitido.", status=405)

    try:
        data = json.loads(request.body)

        nome = str(data.get("nome", "")).strip()
        telefone = str(data.get("telefone", "")).strip()
        cidade = str(data.get("cidade", "")).strip().lower()
        endereco = str(data.get("endereco", "")).strip()
        produto_id = data.get("produto_id")
        quantidade = int(data.get("quantidade", 1))
        forma_pagamento = str(data.get("forma_pagamento", "dinheiro")).strip().lower()
        observacoes = str(data.get("observacoes", "")).strip()
        cupom_codigo = str(data.get("cupom", "")).strip()
        validar_apenas = bool(data.get("validar_apenas", False))

        if not nome:
            return api_error("Nome é obrigatório.", status=400)

        if not telefone:
            return api_error("Telefone é obrigatório.", status=400)

        if not cidade:
            return api_error("Cidade é obrigatória.", status=400)

        if not endereco:
            return api_error("Endereço é obrigatório.", status=400)

        if not produto_id:
            return api_error("Produto é obrigatório.", status=400)

        if quantidade < 1:
            return api_error("Quantidade inválida.", status=400)

        loja = Loja.objects.filter(cidade__iexact=cidade).first()

        if not loja:
            return api_error("Nenhuma loja atende essa cidade.", status=404)

        if not loja_esta_aberta(loja):
            return api_error(
                f"A loja está fechada no momento. Horário de atendimento: {loja.horario_abertura.strftime('%H:%M')} às {loja.horario_fechamento.strftime('%H:%M')}.",
                status=400
            )

        produto = Produto.objects.filter(id=produto_id, loja=loja).first()

        if not produto:
            return api_error("Produto não encontrado para essa loja.", status=404)

        preco_unitario = Decimal(str(produto.preco_venda))
        subtotal = (preco_unitario * Decimal(quantidade)).quantize(Decimal("0.01"))

        resultado_cupom = validar_cupom(cupom_codigo, subtotal)

        if not resultado_cupom["valido"]:
            return api_error(resultado_cupom["erro"], status=400)

        desconto = resultado_cupom["desconto"]
        cupom = resultado_cupom["cupom"]
        cupom_aplicado = cupom.codigo if cupom else ""
        total_final = (subtotal - desconto).quantize(Decimal("0.01"))

        if validar_apenas:
            dados = {
                "cupom_aplicado": cupom_aplicado,
                "desconto": str(desconto),
                "subtotal": str(subtotal),
                "total_final": str(total_final),
            }
            return JsonResponse({
                "ok": True,
                "mensagem": "Cupom validado com sucesso.",
                "dados": dados,
                **dados,
            }, status=200)

        cliente, _ = Cliente.objects.get_or_create(
            telefone=telefone,
            defaults={
                "nome": nome,
                "endereco": endereco,
                "loja": loja,
            }
        )

        cliente.nome = nome
        cliente.endereco = endereco
        cliente.loja = loja
        cliente.save()

        pedido = Pedido.objects.create(
            loja=loja,
            cliente=cliente,
            produto=produto,
            quantidade=quantidade,
            preco_unitario=preco_unitario,
            forma_pagamento=forma_pagamento,
            observacoes=observacoes,
            status="novo",
        )

        if cupom:
            cupom.total_usado += 1
            cupom.save(update_fields=["total_usado"])

        dados = {
            "pedido_id": pedido.id,
            "nome_loja": loja.nome,
            "cidade": loja.cidade,
            "tempo_min": loja.tempo_entrega_min,
            "tempo_max": loja.tempo_entrega_max,
            "subtotal": str(subtotal),
            "desconto": str(desconto),
            "total_final": str(total_final),
            "cupom_aplicado": cupom_aplicado,
            "telefone": telefone,
        }

        return JsonResponse({
            "ok": True,
            "mensagem": "Pedido criado com sucesso.",
            "dados": dados,
            **dados,
        }, status=201)

    except json.JSONDecodeError:
        return api_error("JSON inválido.", status=400)
    except (InvalidOperation, TypeError, ValueError):
        return api_error("Dados inválidos no pedido.", status=400)
    except Exception as e:
        return api_error(str(e), status=500)


@login_required
def pedidos_json(request):
    loja = obter_loja_ativa(request)

    if not loja:
        return JsonResponse({"erro": "Usuário sem loja vinculada."}, status=400)

    pedidos = Pedido.objects.filter(loja=loja).select_related(
        "cliente", "produto"
    ).order_by("-data_pedido")

    lista = []
    for pedido in pedidos:
        lista.append({
            "id": pedido.id,
            "data": pedido.data_pedido.strftime("%d/%m/%Y %H:%M"),
            "cliente": pedido.cliente.nome if pedido.cliente else "Sem cliente",
            "telefone": pedido.cliente.telefone if pedido.cliente else "",
            "endereco": pedido.cliente.endereco if pedido.cliente else "",
            "produto": pedido.produto.nome,
            "quantidade": pedido.quantidade,
            "total": f"{pedido.total_pedido():.2f}",
            "status": pedido.status,
            "status_display": pedido.get_status_display(),
            "observacoes": pedido.observacoes or "",
        })

    return JsonResponse({"pedidos": lista})


def gerar_comissao_frete(pedido, loja):
    """Gera comissão de frete fixo ao entregador quando pedido é marcado como Entregue."""
    entregador = pedido.entregador
    if not entregador or not entregador.recebe_frete:
        return
    if Comissao.objects.filter(pedido=pedido, tipo="frete_fixo").exists():
        return  # evita duplicata
    valor_pedido = (pedido.preco_unitario * pedido.quantidade).quantize(Decimal("0.01"))
    Comissao.objects.create(
        loja=loja,
        entregador=entregador,
        pedido=pedido,
        tipo="frete_fixo",
        valor_venda=valor_pedido,
        percentual=Decimal("0"),
        valor_comissao=entregador.valor_frete,
    )


@login_required
def atualizar_status_pedido(request, pedido_id):
    loja = obter_loja_usuario(request.user)

    if not loja:
        return render(request, "erro_loja.html")

    if request.method != "POST":
        return redirect("pedidos")

    if dia_fechado(loja):
        return render(request, "operacao_bloqueada.html", {
            "mensagem": "O dia de hoje já foi fechado. Não é possível alterar pedidos."
        })

    try:
        pedido = Pedido.objects.get(id=pedido_id, loja=loja)
    except Pedido.DoesNotExist:
        return render(request, "operacao_bloqueada.html", {
            "mensagem": "Pedido não encontrado."
        })

    novo_status = request.POST.get("status")

    status_validos = ["novo", "preparando", "saiu_entrega", "entregue", "cancelado"]

    if novo_status not in status_validos:
        return render(request, "operacao_bloqueada.html", {
            "mensagem": "Status inválido."
        })

    pedido.status = novo_status
    pedido.save()

    if novo_status == "entregue":
        gerar_comissao_frete(pedido, loja)

    registrar_auditoria(
        loja=loja,
        usuario=request.user,
        acao="Status do pedido alterado",
        descricao=f"Pedido #{pedido.id} alterado para {pedido.get_status_display()}"
    )

    messages.success(request, f"Pedido #{pedido.id} atualizado para {pedido.get_status_display()}.")
    return redirect("pedidos")


@login_required
def alterar_status_pedido(request, pedido_id, novo_status):
    loja = obter_loja_usuario(request.user)

    if not loja:
        return render(request, "erro_loja.html")

    if request.method != "POST":
        return redirect("pedidos")

    if dia_fechado(loja):
        return render(request, "operacao_bloqueada.html", {
            "mensagem": "O dia de hoje já foi fechado. Não é possível alterar pedidos."
        })

    pedido = get_object_or_404(Pedido, id=pedido_id, loja=loja)

    status_validos = ["novo", "preparando", "saiu_entrega", "entregue", "cancelado"]

    if novo_status not in status_validos:
        messages.error(request, "Status inválido.")
        return redirect("pedidos")

    pedido.status = novo_status
    pedido.save()

    if novo_status == "entregue":
        gerar_comissao_frete(pedido, loja)

    registrar_auditoria(
        loja=loja,
        usuario=request.user,
        acao="Status do pedido alterado",
        descricao=f"Pedido #{pedido.id} alterado para {pedido.get_status_display()}"
    )

    messages.success(request, f"Pedido #{pedido.id} atualizado para {pedido.get_status_display()}.")
    return redirect("pedidos")


@csrf_exempt
def acompanhar_pedido_site(request):
    pedido_id = request.GET.get("pedido_id", "").strip()
    telefone = request.GET.get("telefone", "").strip()

    if not pedido_id:
        return api_error("Pedido não informado.", status=400)

    if not telefone:
        return api_error("Telefone não informado.", status=400)

    pedido = Pedido.objects.filter(
        id=pedido_id,
        cliente__telefone=telefone
    ).select_related("cliente", "produto", "loja").first()

    if not pedido:
        return api_error("Pedido não encontrado.", status=404)

    subtotal = (Decimal(pedido.preco_unitario) * Decimal(pedido.quantidade)).quantize(Decimal("0.01"))

    pedido_data = {
        "id": pedido.id,
        "status": pedido.status,
        "status_display": pedido.get_status_display(),
        "produto": pedido.produto.nome if pedido.produto else "",
        "quantidade": pedido.quantidade,
        "preco_unitario": str(pedido.preco_unitario),
        "subtotal": str(subtotal),
        "total": str(subtotal),
        "cidade_loja": pedido.loja.cidade if pedido.loja else "",
        "loja": pedido.loja.nome if pedido.loja else "",
        "cliente": pedido.cliente.nome if pedido.cliente else "",
        "telefone": pedido.cliente.telefone if pedido.cliente else "",
        "endereco": pedido.cliente.endereco if pedido.cliente else "",
        "observacoes": pedido.observacoes,
        "data_pedido": pedido.data_pedido.strftime("%d/%m/%Y %H:%M"),
        "tempo_entrega_min": pedido.loja.tempo_entrega_min if pedido.loja else None,
        "tempo_entrega_max": pedido.loja.tempo_entrega_max if pedido.loja else None,
    }

    return JsonResponse({
        "ok": True,
        "mensagem": "Pedido localizado com sucesso.",
        "dados": {
            "pedido": pedido_data,
        },
        # compatibilidade temporária com frontend atual
        "pedido": pedido_data,
    })


@login_required
def fechamento_caixa(request):
    loja = obter_loja_usuario(request.user)

    if not loja:
        return render(request, "erro_loja.html")

    if not usuario_eh_gerente_ou_admin(request.user):
        return render(request, "operacao_bloqueada.html", {
            "mensagem": "Apenas gerente ou admin podem acessar o fechamento de caixa."
        })

    hoje = timezone.now().date()

    vendas = Venda.objects.filter(
        data_venda__date=hoje,
        loja=loja,
        status="ativa"
    )

    pedidos = Pedido.objects.filter(
        data_pedido__date=hoje,
        loja=loja
    ).exclude(status="cancelado")

    total_vendas = 0
    for venda in vendas:
        total_vendas += float(venda.quantidade * venda.preco_unitario)

    total_pedidos = 0
    for pedido in pedidos:
        total_pedidos += float(pedido.quantidade * pedido.preco_unitario)

    dinheiro_vendas, pix_vendas, credito_vendas, debito_vendas = somar_pagamentos_mistos(vendas)
    dinheiro_pedidos, pix_pedidos, credito_pedidos, debito_pedidos = somar_pagamentos_pedidos(pedidos)

    return render(request, "fechamento_caixa.html", {
        "loja": loja,
        "data": hoje,
        "quantidade_vendas": vendas.count(),
        "quantidade_pedidos": pedidos.count(),
        "total_vendas": total_vendas,
        "total_pedidos": total_pedidos,
        "dinheiro_vendas": dinheiro_vendas,
        "pix_vendas": pix_vendas,
        "credito_vendas": credito_vendas,
        "debito_vendas": debito_vendas,
        "dinheiro_pedidos": dinheiro_pedidos,
        "pix_pedidos": pix_pedidos,
        "credito_pedidos": credito_pedidos,
        "debito_pedidos": debito_pedidos,
        "total_geral": total_vendas + total_pedidos,
    })

@login_required
def cancelar_venda(request, venda_id):
    loja = obter_loja_usuario(request.user)

    if not loja:
        return render(request, "erro_loja.html")

    if dia_fechado(loja):
        return render(request, "operacao_bloqueada.html", {
            "mensagem": "O dia de hoje já foi fechado. Não é possível cancelar vendas."
        })

    try:
        venda = Venda.objects.get(id=venda_id, loja=loja)
    except Venda.DoesNotExist:
        return render(request, "operacao_bloqueada.html", {
            "mensagem": "Venda não encontrada."
        })

    if venda.status == "cancelada":
        return render(request, "operacao_bloqueada.html", {
            "mensagem": "Essa venda já foi cancelada."
        })

    if request.method != "POST":
        cliente_nome = venda.cliente.nome if venda.cliente else "Sem cliente"
        return render(request, "confirmar_acao.html", {
            "titulo": "Cancelar venda",
            "mensagem": "Tem certeza que deseja cancelar esta venda? Essa ação altera o estoque.",
            "detalhes": f"Venda #{venda.id} | Cliente: {cliente_nome} | Produto: {venda.produto.nome} | Quantidade: {venda.quantidade}",
            "voltar_url": "/venda/",
        })

    produto = venda.produto
    quantidade = venda.quantidade

    if not produto.controla_retorno:
        produto.estoque_cheio += quantidade
    else:
        if venda.tipo_venda == "troca":
            produto.estoque_cheio += quantidade
            produto.estoque_vazio -= quantidade
            if produto.estoque_vazio < 0:
                produto.estoque_vazio = 0
        elif venda.tipo_venda == "completo":
            produto.estoque_cheio += quantidade
        elif venda.tipo_venda == "casco":
            produto.estoque_vazio += quantidade
        else:
            produto.estoque_cheio += quantidade

    produto.save()

    venda.status = "cancelada"
    venda.save()

    MovimentacaoEstoque.objects.create(
        loja=loja,
        produto=produto,
        usuario=request.user,
        tipo="entrada",
        quantidade=quantidade,
        motivo=f"Cancelamento da venda #{venda.id}"
    )

    registrar_auditoria(
        loja=loja,
        usuario=request.user,
        acao="Venda cancelada",
        descricao=f"Venda #{venda.id} | Produto: {produto.nome} | Quantidade: {quantidade}"
    )

    return redirect("/venda/")

@login_required
def novo_cliente(request):
    loja = obter_loja_usuario(request.user)

    if not loja:
        return render(request, "erro_loja.html")

    if request.method == "POST":
        nome = request.POST.get("nome", "").strip()
        telefone = request.POST.get("telefone", "").strip()
        endereco = request.POST.get("endereco", "").strip()
        observacoes = request.POST.get("observacoes", "").strip()
        observacao_comercial = request.POST.get("observacao_comercial", "").strip()

        if not nome:
            return render(request, "novo_cliente.html", {
                "loja": loja,
                "erro": "O nome do cliente é obrigatório.",
            })

        Cliente.objects.create(
            loja=loja,
            nome=nome,
            telefone=telefone,
            endereco=endereco,
            observacoes=observacoes,
            observacao_comercial=observacao_comercial,
            ativo=True,
        )

        return redirect("lista_clientes")

    return render(request, "novo_cliente.html", {"loja": loja})


@login_required
def criar_cliente(request):
    return novo_cliente(request)



def montar_dados_historico_cliente(request, cliente_id):
    loja = obter_loja_usuario(request.user)

    if not loja:
        return {"erro_loja": True}

    try:
        cliente = Cliente.objects.get(id=cliente_id, loja=loja)
    except Cliente.DoesNotExist:
        return {"cliente_nao_encontrado": True}

    hoje = timezone.now().date()

    filtro_periodo = request.GET.get("filtro", "30dias").strip()
    data_inicial_str = request.GET.get("data_inicial", "").strip()
    data_final_str = request.GET.get("data_final", "").strip()

    busca = request.GET.get("busca", "").strip()
    secao = request.GET.get("secao", "").strip()

    status_pedido = request.GET.get("status_pedido", "").strip()
    tipo_venda = request.GET.get("tipo_venda", "").strip()
    status_venda = request.GET.get("status_venda", "").strip()

    data_inicial = hoje - timedelta(days=29)
    data_final = hoje

    if filtro_periodo == "hoje":
        data_inicial = hoje
        data_final = hoje
    elif filtro_periodo == "7dias":
        data_inicial = hoje - timedelta(days=6)
        data_final = hoje
    elif filtro_periodo == "30dias":
        data_inicial = hoje - timedelta(days=29)
        data_final = hoje
    elif filtro_periodo == "mes":
        data_inicial = hoje.replace(day=1)
        data_final = hoje
    elif filtro_periodo == "personalizado" and data_inicial_str and data_final_str:
        try:
            data_inicial = datetime.strptime(data_inicial_str, "%Y-%m-%d").date()
            data_final = datetime.strptime(data_final_str, "%Y-%m-%d").date()
        except ValueError:
            data_inicial = hoje - timedelta(days=29)
            data_final = hoje

    if data_inicial > data_final:
        data_inicial, data_final = data_final, data_inicial

    vendas_qs = Venda.objects.filter(
        loja=loja,
        cliente=cliente,
        data_venda__date__gte=data_inicial,
        data_venda__date__lte=data_final,
    ).select_related("produto", "funcionario", "cliente")

    pedidos_qs = Pedido.objects.filter(
        loja=loja,
        cliente=cliente,
        data_pedido__date__gte=data_inicial,
        data_pedido__date__lte=data_final,
    ).select_related("cliente", "produto", "entregador")

    if tipo_venda:
        vendas_qs = vendas_qs.filter(tipo_venda=tipo_venda)
    if status_venda:
        vendas_qs = vendas_qs.filter(status=status_venda)
    if status_pedido:
        pedidos_qs = pedidos_qs.filter(status=status_pedido)

    if busca:
        vendas_qs = vendas_qs.filter(
            models.Q(produto__nome__icontains=busca) |
            models.Q(funcionario__username__icontains=busca)
        )
        pedidos_qs = pedidos_qs.filter(
            models.Q(produto__nome__icontains=busca) |
            models.Q(observacoes__icontains=busca) |
            models.Q(status__icontains=busca)
        )

    if secao == "vendas":
        pedidos_qs = Pedido.objects.none()
    elif secao == "pedidos":
        vendas_qs = Venda.objects.none()

    vendas_qs = vendas_qs.order_by("-data_venda")
    pedidos_qs = pedidos_qs.order_by("-data_pedido")

    total_vendas = vendas_qs.count()
    total_pedidos = pedidos_qs.count()
    quantidade_vendida = sum(item.quantidade for item in vendas_qs)
    valor_total_vendas = sum(item.total_venda() for item in vendas_qs)

    venda_page_num = request.GET.get("venda_page", 1)
    pedido_page_num = request.GET.get("pedido_page", 1)

    vendas = Paginator(vendas_qs, 15).get_page(venda_page_num)
    pedidos = Paginator(pedidos_qs, 15).get_page(pedido_page_num)

    params = request.GET.copy()
    for chave in ["venda_page", "pedido_page"]:
        if chave in params:
            params.pop(chave)

    querystring_base = params.urlencode()

    return {
        "loja": loja,
        "cliente": cliente,
        "vendas": vendas,
        "pedidos": pedidos,
        "filtro": filtro_periodo,
        "data_inicial": data_inicial,
        "data_final": data_final,
        "busca": busca,
        "secao": secao,
        "status_pedido": status_pedido,
        "tipo_venda": tipo_venda,
        "status_venda": status_venda,
        "total_vendas": total_vendas,
        "total_pedidos": total_pedidos,
        "quantidade_vendida": quantidade_vendida,
        "valor_total_vendas": valor_total_vendas,
        "tipos_venda": Venda.TIPO_VENDA_CHOICES,
        "status_venda_choices": Venda.STATUS_CHOICES,
        "status_pedido_choices": Pedido.STATUS_CHOICES,
        "querystring_base": querystring_base,
    }


@login_required
def historico_cliente(request, cliente_id):
    contexto = montar_dados_historico_cliente(request, cliente_id)

    if contexto.get("erro_loja"):
        return render(request, "erro_loja.html")

    if contexto.get("cliente_nao_encontrado"):
        return render(request, "operacao_bloqueada.html", {
            "mensagem": "Cliente não encontrado."
        })

    return render(request, "historico_cliente.html", contexto)


@login_required
def historico_cliente_excel(request, cliente_id):
    contexto = montar_dados_historico_cliente(request, cliente_id)

    if contexto.get("erro_loja"):
        return render(request, "erro_loja.html")

    if contexto.get("cliente_nao_encontrado"):
        return render(request, "operacao_bloqueada.html", {
            "mensagem": "Cliente não encontrado."
        })

    loja = contexto["loja"]
    cliente = contexto["cliente"]

    hoje = timezone.now().date()
    filtro_periodo = request.GET.get("filtro", "30dias").strip()
    data_inicial_str = request.GET.get("data_inicial", "").strip()
    data_final_str = request.GET.get("data_final", "").strip()
    busca = request.GET.get("busca", "").strip()
    secao = request.GET.get("secao", "").strip()
    status_pedido = request.GET.get("status_pedido", "").strip()
    tipo_venda = request.GET.get("tipo_venda", "").strip()
    status_venda = request.GET.get("status_venda", "").strip()

    data_inicial = hoje - timedelta(days=29)
    data_final = hoje
    if filtro_periodo == "hoje":
        data_inicial = hoje
        data_final = hoje
    elif filtro_periodo == "7dias":
        data_inicial = hoje - timedelta(days=6)
    elif filtro_periodo == "mes":
        data_inicial = hoje.replace(day=1)
    elif filtro_periodo == "personalizado" and data_inicial_str and data_final_str:
        try:
            data_inicial = datetime.strptime(data_inicial_str, "%Y-%m-%d").date()
            data_final = datetime.strptime(data_final_str, "%Y-%m-%d").date()
        except ValueError:
            pass

    vendas_qs = Venda.objects.filter(
        loja=loja,
        cliente=cliente,
        data_venda__date__gte=data_inicial,
        data_venda__date__lte=data_final,
    ).select_related("produto", "funcionario", "cliente")

    pedidos_qs = Pedido.objects.filter(
        loja=loja,
        cliente=cliente,
        data_pedido__date__gte=data_inicial,
        data_pedido__date__lte=data_final,
    ).select_related("cliente", "produto", "entregador")

    if tipo_venda:
        vendas_qs = vendas_qs.filter(tipo_venda=tipo_venda)
    if status_venda:
        vendas_qs = vendas_qs.filter(status=status_venda)
    if status_pedido:
        pedidos_qs = pedidos_qs.filter(status=status_pedido)
    if busca:
        vendas_qs = vendas_qs.filter(
            models.Q(produto__nome__icontains=busca) |
            models.Q(funcionario__username__icontains=busca)
        )
        pedidos_qs = pedidos_qs.filter(
            models.Q(produto__nome__icontains=busca) |
            models.Q(observacoes__icontains=busca) |
            models.Q(status__icontains=busca)
        )
    if secao == "vendas":
        pedidos_qs = Pedido.objects.none()
    elif secao == "pedidos":
        vendas_qs = Venda.objects.none()

    vendas_qs = vendas_qs.order_by("-data_venda")
    pedidos_qs = pedidos_qs.order_by("-data_pedido")

    total_vendas = vendas_qs.count()
    total_pedidos = pedidos_qs.count()
    quantidade_vendida = sum(item.quantidade for item in vendas_qs)
    valor_total_vendas = sum(item.total_venda() for item in vendas_qs)

    wb = Workbook()
    negrito = Font(bold=True)
    ws_resumo = wb.active
    ws_resumo.title = "Resumo"

    for linha in [
        ["Loja", loja.nome],
        ["Cliente", cliente.nome],
        ["Filtro", filtro_periodo],
        ["Data inicial", data_inicial.strftime("%d/%m/%Y")],
        ["Data final", data_final.strftime("%d/%m/%Y")],
        ["Total de vendas", total_vendas],
        ["Total de pedidos", total_pedidos],
        ["Quantidade vendida", quantidade_vendida],
        ["Valor total vendas", float(valor_total_vendas or 0)],
        ["Busca", busca or "-"],
        ["Seção", secao or "Tudo"],
        ["Status pedido", status_pedido or "-"],
        ["Tipo venda", tipo_venda or "-"],
        ["Status venda", status_venda or "-"],
    ]:
        ws_resumo.append(linha)
    for cell in ws_resumo[1]:
        cell.font = negrito

    ws_vendas = wb.create_sheet("Vendas")
    ws_vendas.append(["Data", "Produto", "Quantidade", "Tipo de venda", "Status", "Funcionario", "Preco unitario", "Total"])
    for cell in ws_vendas[1]:
        cell.font = negrito
    for venda in vendas_qs:
        ws_vendas.append([
            venda.data_venda.strftime("%d/%m/%Y %H:%M"),
            venda.produto.nome if venda.produto else "",
            venda.quantidade,
            venda.get_tipo_venda_display(),
            venda.get_status_display(),
            venda.funcionario.username if venda.funcionario else "",
            float(venda.preco_unitario or 0),
            float(venda.total_venda() or 0),
        ])

    ws_pedidos = wb.create_sheet("Pedidos")
    ws_pedidos.append(["Data", "Produto", "Quantidade", "Status", "Entregador", "Observacoes"])
    for cell in ws_pedidos[1]:
        cell.font = negrito
    for pedido in pedidos_qs:
        ws_pedidos.append([
            pedido.data_pedido.strftime("%d/%m/%Y %H:%M"),
            pedido.produto.nome if pedido.produto else "",
            pedido.quantidade,
            pedido.get_status_display(),
            pedido.entregador.nome if pedido.entregador else "",
            pedido.observacoes or "",
        ])

    for ws in wb.worksheets:
        for coluna in ws.columns:
            maior = 0
            letra = coluna[0].column_letter
            for celula in coluna:
                valor = str(celula.value) if celula.value is not None else ""
                maior = max(maior, len(valor))
            ws.column_dimensions[letra].width = min(maior + 2, 40)

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="historico_cliente_{cliente.id}.xlsx"'
    wb.save(response)
    return response


@login_required
def lista_clientes(request):
    loja = obter_loja_usuario(request.user)

    if not loja:
        return render(request, "erro_loja.html")

    busca = request.GET.get("q", "").strip()
    pagina_atual = request.GET.get("page", 1)
    mostrar_inativos = request.GET.get("inativos") == "1"

    clientes_qs = Cliente.objects.filter(loja=loja)
    if not mostrar_inativos:
        clientes_qs = clientes_qs.filter(ativo=True)

    if busca:
        clientes_qs = clientes_qs.filter(
            models.Q(nome__icontains=busca) |
            models.Q(telefone__icontains=busca) |
            models.Q(endereco__icontains=busca) |
            models.Q(observacoes__icontains=busca) |
            models.Q(observacao_comercial__icontains=busca)
        )

    clientes_qs = clientes_qs.order_by("nome")
    total_clientes = clientes_qs.count()
    paginator = Paginator(clientes_qs, 15)
    clientes = paginator.get_page(pagina_atual)

    params = request.GET.copy()
    if "page" in params:
        params.pop("page")
    querystring_base = params.urlencode()

    return render(request, "lista_clientes.html", {
        "loja": loja,
        "clientes": clientes,
        "busca": busca,
        "total_clientes": total_clientes,
        "querystring_base": querystring_base,
        "mostrar_inativos": mostrar_inativos,
    })


@login_required
def clientes_excel(request):
    loja = obter_loja_usuario(request.user)

    if not loja:
        return render(request, "erro_loja.html")

    busca = request.GET.get("q", "").strip()
    mostrar_inativos = request.GET.get("inativos") == "1"

    clientes = Cliente.objects.filter(loja=loja)
    if not mostrar_inativos:
        clientes = clientes.filter(ativo=True)

    if busca:
        clientes = clientes.filter(
            models.Q(nome__icontains=busca) |
            models.Q(telefone__icontains=busca) |
            models.Q(endereco__icontains=busca) |
            models.Q(observacoes__icontains=busca) |
            models.Q(observacao_comercial__icontains=busca)
        )

    clientes = clientes.order_by("nome")

    wb = Workbook()
    ws = wb.active
    ws.title = "Clientes"
    negrito = Font(bold=True)
    ws.append(["ID", "Nome", "Telefone", "Endereço", "Observações", "Observação comercial", "Último contato", "Contatado hoje", "Ativo", "Criado em"])
    for cell in ws[1]:
        cell.font = negrito

    for cliente in clientes:
        ws.append([
            cliente.id,
            cliente.nome,
            cliente.telefone or "",
            cliente.endereco or "",
            cliente.observacoes or "",
            cliente.observacao_comercial or "",
            cliente.ultimo_contato.strftime("%d/%m/%Y") if cliente.ultimo_contato else "",
            "Sim" if cliente.contatado_hoje else "Não",
            "Sim" if cliente.ativo else "Não",
            cliente.criado_em.strftime("%d/%m/%Y %H:%M") if cliente.criado_em else "",
        ])

    for coluna in ws.columns:
        maior = 0
        letra = coluna[0].column_letter
        for celula in coluna:
            valor = str(celula.value) if celula.value is not None else ""
            maior = max(maior, len(valor))
        ws.column_dimensions[letra].width = min(maior + 2, 40)

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="clientes.xlsx"'
    wb.save(response)
    return response


@login_required
def editar_cliente(request, cliente_id):
    loja = obter_loja_usuario(request.user)

    if not loja:
        return render(request, "erro_loja.html")

    try:
        cliente = Cliente.objects.get(id=cliente_id, loja=loja)
    except Cliente.DoesNotExist:
        return render(request, "operacao_bloqueada.html", {"mensagem": "Cliente não encontrado."})

    if request.method == "POST":
        nome = request.POST.get("nome", "").strip()
        if not nome:
            return render(request, "editar_cliente.html", {"loja": loja, "cliente": cliente, "erro": "O nome do cliente é obrigatório."})

        cliente.nome = nome
        cliente.telefone = request.POST.get("telefone", "").strip()
        cliente.endereco = request.POST.get("endereco", "").strip()
        cliente.observacoes = request.POST.get("observacoes", "").strip()
        cliente.observacao_comercial = request.POST.get("observacao_comercial", "").strip()

        ultimo_contato = request.POST.get("ultimo_contato", "").strip()
        if ultimo_contato:
            try:
                cliente.ultimo_contato = datetime.strptime(ultimo_contato, "%Y-%m-%d").date()
            except ValueError:
                pass
        else:
            cliente.ultimo_contato = None

        cliente.contatado_hoje = request.POST.get("contatado_hoje") == "on"
        cliente.save()
        return redirect("lista_clientes")

    return render(request, "editar_cliente.html", {"loja": loja, "cliente": cliente})


@login_required
def inativar_cliente(request, cliente_id):
    loja = obter_loja_usuario(request.user)

    if not loja:
        return render(request, "erro_loja.html")

    try:
        cliente = Cliente.objects.get(id=cliente_id, loja=loja)
    except Cliente.DoesNotExist:
        return render(request, "operacao_bloqueada.html", {"mensagem": "Cliente não encontrado."})

    if request.method == "POST":
        cliente.ativo = False
        cliente.save()
        return redirect("lista_clientes")

    return render(request, "confirmar_acao.html", {
        "titulo": "Inativar cliente",
        "mensagem": f"Tem certeza que deseja inativar o cliente {cliente.nome}?",
        "botao": "Inativar cliente",
    })


@login_required
def clientes_recompra(request):
    loja = obter_loja_usuario(request.user)

    if not loja:
        return render(request, "erro_loja.html")

    hoje = timezone.now().date()
    clientes_lista = []

    clientes = Cliente.objects.filter(loja=loja, ativo=True).order_by("nome")

    for cliente in clientes:
        ultima_venda = Venda.objects.filter(
            loja=loja,
            cliente=cliente,
            status="ativa"
        ).order_by("-data_venda").first()

        if ultima_venda:
            ultima_compra = ultima_venda.data_venda.date()
            dias_sem_comprar = (hoje - ultima_compra).days
            if dias_sem_comprar >= 20:
                clientes_lista.append({
                    "cliente": cliente,
                    "ultima_compra": ultima_compra,
                    "dias_sem_comprar": dias_sem_comprar,
                })

    clientes_lista = sorted(clientes_lista, key=lambda x: x["dias_sem_comprar"], reverse=True)

    return render(request, "clientes_recompra.html", {"loja": loja, "clientes": clientes_lista})


@login_required
def marcar_contato_cliente(request, cliente_id):
    loja = obter_loja_usuario(request.user)

    if not loja:
        return render(request, "erro_loja.html")

    cliente = Cliente.objects.get(id=cliente_id, loja=loja)

    if request.method == "POST":
        observacao = request.POST.get("observacao_comercial", "")
        cliente.ultimo_contato = timezone.now().date()
        cliente.contatado_hoje = True
        cliente.observacao_comercial = observacao
        cliente.save()

    return redirect("/recompras/")


@login_required
def relatorio_entregadores(request):
    loja = obter_loja_usuario(request.user)

    if not loja:
        return render(request, "erro_loja.html")

    hoje = timezone.now().date()
    filtro = request.GET.get("filtro", "hoje")

    data_inicial = hoje
    data_final = hoje

    if filtro == "ontem":
        data_inicial = hoje - timedelta(days=1)
        data_final = data_inicial

    elif filtro == "7dias":
        data_inicial = hoje - timedelta(days=6)
        data_final = hoje

    elif filtro == "mes":
        data_inicial = hoje.replace(day=1)
        data_final = hoje

    elif filtro == "personalizado":
        data_ini_str = request.GET.get("data_inicial")
        data_fim_str = request.GET.get("data_final")

        if data_ini_str and data_fim_str:
            data_inicial = datetime.strptime(data_ini_str, "%Y-%m-%d").date()
            data_final = datetime.strptime(data_fim_str, "%Y-%m-%d").date()

    entregadores = Entregador.objects.filter(loja=loja).order_by("nome")

    dados_entregadores = []

    for entregador in entregadores:
        pedidos = Pedido.objects.filter(
            loja=loja,
            entregador=entregador,
            data_pedido__date__gte=data_inicial,
            data_pedido__date__lte=data_final
        )

        total_pedidos = pedidos.count()
        pedidos_entregues = pedidos.filter(status="entregue").count()
        pedidos_em_rota = pedidos.filter(status="saiu_entrega").count()
        pedidos_preparando = pedidos.filter(status="preparando").count()
        pedidos_novos = pedidos.filter(status="novo").count()

        valor_total = 0
        for pedido in pedidos:
            valor_total += pedido.quantidade * pedido.preco_unitario

        dados_entregadores.append({
            "entregador": entregador,
            "total_pedidos": total_pedidos,
            "pedidos_entregues": pedidos_entregues,
            "pedidos_em_rota": pedidos_em_rota,
            "pedidos_preparando": pedidos_preparando,
            "pedidos_novos": pedidos_novos,
            "valor_total": valor_total,
        })

    return render(request, "relatorio_entregadores.html", {
        "loja": loja,
        "dados_entregadores": dados_entregadores,
        "data_inicial": data_inicial,
        "data_final": data_final,
        "filtro": filtro,
    })


@login_required
def painel_entregador(request):
    loja = obter_loja_usuario(request.user)

    if not loja:
        return render(request, "erro_loja.html")

    pedidos_em_rota = Pedido.objects.filter(
        loja=loja,
        status="saiu_entrega"
    ).order_by("-data_pedido")

    return render(request, "painel_entregador.html", {
        "loja": loja,
        "pedidos": pedidos_em_rota,
    })

@login_required
def entregar_pedido(request, pedido_id):
    loja = obter_loja_usuario(request.user)

    if not loja:
        return render(request, "erro_loja.html")

    if request.method != "POST":
        return redirect("/painel-entregador/")

    try:
        pedido = Pedido.objects.get(id=pedido_id, loja=loja)
    except Pedido.DoesNotExist:
        return render(request, "operacao_bloqueada.html", {
            "mensagem": "Pedido não encontrado."
        })

    pedido.status = "entregue"
    pedido.save()

    gerar_comissao_frete(pedido, loja)

    return redirect("/painel-entregador/")

@login_required
def fechamento_caixa_pdf(request):
    loja = obter_loja_usuario(request.user)

    if not loja:
        return render(request, "erro_loja.html")

    if not usuario_eh_gerente_ou_admin(request.user):
        return render(request, "operacao_bloqueada.html", {
            "mensagem": "Apenas gerente ou admin podem exportar o fechamento."
        })

    hoje = timezone.now().date()

    vendas = Venda.objects.filter(
        data_venda__date=hoje,
        loja=loja,
        status="ativa"
    )

    pedidos = Pedido.objects.filter(
        data_pedido__date=hoje,
        loja=loja
    ).exclude(status="cancelado")

    despesas = Despesa.objects.filter(loja=loja, data__date=hoje)
    retiradas = RetiradaFuncionario.objects.filter(loja=loja, data__date=hoje)

    total_vendas = sum(float(v.quantidade * v.preco_unitario) for v in vendas)
    total_pedidos = sum(float(p.quantidade * p.preco_unitario) for p in pedidos)
    total_despesas = sum(float(d.valor) for d in despesas)
    total_retiradas = sum(float(r.valor) for r in retiradas)

    dinheiro_vendas, pix_vendas, credito_vendas, debito_vendas = somar_pagamentos_mistos(vendas)
    dinheiro_pedidos, pix_pedidos, credito_pedidos, debito_pedidos = somar_pagamentos_pedidos(pedidos)

    total_geral = total_vendas + total_pedidos - total_despesas - total_retiradas

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'inline; filename="fechamento_caixa.pdf"'

    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas

    p = canvas.Canvas(response, pagesize=A4)
    largura, altura = A4

    y = altura - 50
    p.setFont("Helvetica-Bold", 16)
    p.drawString(50, y, f"Fechamento de Caixa - {loja.nome}")
    y -= 25
    p.setFont("Helvetica", 11)
    p.drawString(50, y, f"Data: {hoje.strftime('%d/%m/%Y')}")
    y -= 30

    linhas = [
        f"Total de vendas: R$ {total_vendas:.2f}",
        f"Total de pedidos: R$ {total_pedidos:.2f}",
        f"Total de despesas: R$ {total_despesas:.2f}",
        f"Total de retiradas: R$ {total_retiradas:.2f}",
        f"Total geral: R$ {total_geral:.2f}",
        "",
        "Pagamentos - Vendas",
        f"Dinheiro: R$ {dinheiro_vendas:.2f}",
        f"PIX: R$ {pix_vendas:.2f}",
        f"Crédito: R$ {credito_vendas:.2f}",
        f"Débito: R$ {debito_vendas:.2f}",
        "",
        "Pagamentos - Pedidos",
        f"Dinheiro: R$ {dinheiro_pedidos:.2f}",
        f"PIX: R$ {pix_pedidos:.2f}",
        f"Crédito: R$ {credito_pedidos:.2f}",
        f"Débito: R$ {debito_pedidos:.2f}",
    ]

    for linha in linhas:
        p.drawString(50, y, linha)
        y -= 18
        if y < 50:
            p.showPage()
            y = altura - 50
            p.setFont("Helvetica", 11)

    p.showPage()
    p.save()
    return response

@login_required
def fechamento_caixa_excel(request):
    loja = obter_loja_usuario(request.user)

    if not loja:
        return render(request, "erro_loja.html")

    if not usuario_eh_gerente_ou_admin(request.user):
        return render(request, "operacao_bloqueada.html", {
            "mensagem": "Apenas gerente ou admin podem exportar o fechamento."
        })

    hoje = timezone.now().date()

    vendas = Venda.objects.filter(
        data_venda__date=hoje,
        loja=loja,
        status="ativa"
    )

    pedidos = Pedido.objects.filter(
        data_pedido__date=hoje,
        loja=loja
    ).exclude(status="cancelado")

    despesas = Despesa.objects.filter(loja=loja, data__date=hoje)
    retiradas = RetiradaFuncionario.objects.filter(loja=loja, data__date=hoje)

    total_vendas = sum(float(v.quantidade * v.preco_unitario) for v in vendas)
    total_pedidos = sum(float(p.quantidade * p.preco_unitario) for p in pedidos)
    total_despesas = sum(float(d.valor) for d in despesas)
    total_retiradas = sum(float(r.valor) for r in retiradas)

    dinheiro_vendas, pix_vendas, credito_vendas, debito_vendas = somar_pagamentos_mistos(vendas)
    dinheiro_pedidos, pix_pedidos, credito_pedidos, debito_pedidos = somar_pagamentos_pedidos(pedidos)

    total_geral = total_vendas + total_pedidos - total_despesas - total_retiradas

    wb = Workbook()

    ws_resumo = wb.active
    ws_resumo.title = "Resumo"

    ws_resumo.append(["Loja", loja.nome])
    ws_resumo.append(["Data", hoje.strftime("%d/%m/%Y")])
    ws_resumo.append([])
    ws_resumo.append(["Total de vendas", total_vendas])
    ws_resumo.append(["Total de pedidos", total_pedidos])
    ws_resumo.append(["Total de despesas", total_despesas])
    ws_resumo.append(["Total de retiradas", total_retiradas])
    ws_resumo.append(["Total geral", total_geral])
    ws_resumo.append([])
    ws_resumo.append(["Pagamentos vendas", ""])
    ws_resumo.append(["Dinheiro", dinheiro_vendas])
    ws_resumo.append(["PIX", pix_vendas])
    ws_resumo.append(["Crédito", credito_vendas])
    ws_resumo.append(["Débito", debito_vendas])
    ws_resumo.append([])
    ws_resumo.append(["Pagamentos pedidos", ""])
    ws_resumo.append(["Dinheiro", dinheiro_pedidos])
    ws_resumo.append(["PIX", pix_pedidos])
    ws_resumo.append(["Crédito", credito_pedidos])
    ws_resumo.append(["Débito", debito_pedidos])

    ws_vendas = wb.create_sheet("Vendas")
    ws_vendas.append([
        "ID", "Data", "Hora", "Cliente", "Produto", "Quantidade",
        "Preço Unitário", "Pagamento 1", "Valor 1",
        "Pagamento 2", "Valor 2", "Tipo de Venda", "Status", "Funcionário", "Total"
    ])

    for venda in vendas:
        cliente_nome = venda.cliente.nome if venda.cliente else "Sem cliente"
        ws_vendas.append([
            venda.id,
            venda.data_venda.strftime("%d/%m/%Y"),
            venda.data_venda.strftime("%H:%M"),
            cliente_nome,
            venda.produto.nome,
            venda.quantidade,
            float(venda.preco_unitario),
            venda.forma_pagamento_1,
            float(venda.valor_pagamento_1 or 0),
            venda.forma_pagamento_2 or "",
            float(venda.valor_pagamento_2 or 0),
            venda.get_tipo_venda_display(),
            venda.status,
            venda.funcionario.username,
            float(venda.quantidade * venda.preco_unitario),
        ])

    ws_pedidos = wb.create_sheet("Pedidos")
    ws_pedidos.append([
        "ID", "Data", "Cliente", "Produto", "Quantidade",
        "Preço Unitário", "Pagamento", "Status", "Total"
    ])

    for pedido in pedidos:
        cliente_nome = pedido.cliente.nome if pedido.cliente else "Sem cliente"
        ws_pedidos.append([
            pedido.id,
            pedido.data_pedido.strftime("%d/%m/%Y %H:%M"),
            cliente_nome,
            pedido.produto.nome,
            pedido.quantidade,
            float(pedido.preco_unitario),
            pedido.get_forma_pagamento_display(),
            pedido.status,
            float(pedido.quantidade * pedido.preco_unitario),
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="fechamento_caixa.xlsx"'
    wb.save(response)
    return response

@login_required
def despesas(request):
    loja = obter_loja_usuario(request.user)

    if not loja:
        return render(request, "erro_loja.html")

    if not usuario_eh_gerente_ou_admin(request.user):
        return render(request, "operacao_bloqueada.html", {
            "mensagem": "Apenas gerente ou admin podem lançar despesas."
        })

    if dia_fechado(loja):
        return render(request, "operacao_bloqueada.html", {
            "mensagem": "O dia de hoje já foi fechado. Não é possível lançar despesas."
        })

    if request.method == "POST":
        valor = request.POST.get("valor")
        categoria = request.POST.get("categoria")
        descricao = request.POST.get("descricao")

        despesa = Despesa.objects.create(
            loja=loja,
            funcionario=request.user,
            valor=valor,
            categoria=categoria,
            descricao=descricao,
        )

        registrar_auditoria(
            loja=loja,
            usuario=request.user,
            acao="Despesa lançada",
            descricao=f"Despesa #{despesa.id} | Categoria: {categoria} | Valor: R$ {valor}"
        )

        return redirect("/despesas/")

    hoje = timezone.now().date()

    despesas_lista = Despesa.objects.filter(
        loja=loja,
        data__date=hoje
    ).order_by("-data")

    total = 0
    for d in despesas_lista:
        total += float(d.valor)

    return render(request, "despesas.html", {
        "loja": loja,
        "despesas": despesas_lista,
        "total": total,
    })

@login_required
def retiradas_funcionarios(request):
    loja = obter_loja_usuario(request.user)

    if not loja:
        return render(request, "erro_loja.html")

    if not usuario_eh_gerente_ou_admin(request.user):
        return render(request, "operacao_bloqueada.html", {
            "mensagem": "Apenas gerente ou admin podem lançar retiradas."
        })

    if dia_fechado(loja):
        return render(request, "operacao_bloqueada.html", {
            "mensagem": "O dia de hoje já foi fechado. Não é possível lançar retiradas."
        })

    funcionarios = User.objects.filter(
        perfilusuario__loja=loja
    ).order_by("username").distinct()

    if request.method == "POST":
        funcionario_id = request.POST.get("funcionario")
        valor = request.POST.get("valor")
        tipo = request.POST.get("tipo")
        descricao = request.POST.get("descricao")

        funcionario = User.objects.get(id=funcionario_id)

        retirada = RetiradaFuncionario.objects.create(
           loja=loja,
           funcionario=funcionario,
           registrado_por=request.user,
           valor=valor,
           tipo=tipo,
           descricao=descricao,
        )

        registrar_auditoria(
           loja=loja,
           usuario=request.user,
           acao="Retirada lançada",
           descricao=f"Retirada #{retirada.id} | Funcionário: {funcionario.username} | Tipo: {tipo} | Valor: R$ {valor}"
        )

        return redirect("/retiradas-funcionarios/")

    hoje = timezone.now().date()

    retiradas = RetiradaFuncionario.objects.filter(
        loja=loja,
        data__date=hoje
    ).order_by("-data")

    total = 0
    for r in retiradas:
        total += float(r.valor)

    return render(request, "retiradas_funcionarios.html", {
        "loja": loja,
        "funcionarios": funcionarios,
        "retiradas": retiradas,
        "total": total,
    })

@login_required
def lucro_diario(request):
    loja = obter_loja_usuario(request.user)

    if not loja:
        return render(request, "erro_loja.html")

    if not usuario_eh_admin(request.user):
        return render(request, "operacao_bloqueada.html", {
            "mensagem": "Apenas administradores podem acessar o lucro diário."
        })

    hoje = timezone.now().date()

    vendas = Venda.objects.filter(
        loja=loja,
        data_venda__date=hoje,
        status="ativa"
    ).select_related("produto")

    pedidos = Pedido.objects.filter(
        loja=loja,
        data_pedido__date=hoje
    ).exclude(status="cancelado").select_related("produto")

    despesas = Despesa.objects.filter(
        loja=loja,
        data__date=hoje
    )

    retiradas = RetiradaFuncionario.objects.filter(
        loja=loja,
        data__date=hoje
    )

    total_vendas = 0
    lucro_vendas = 0

    for venda in vendas:
        faturamento = float(venda.quantidade * venda.preco_unitario)
        custo = float(venda.quantidade * venda.produto.custo_unitario)

        total_vendas += faturamento
        lucro_vendas += (faturamento - custo)

    total_pedidos = 0
    lucro_pedidos = 0

    for pedido in pedidos:
        faturamento = float(pedido.quantidade * pedido.preco_unitario)
        custo = float(pedido.quantidade * pedido.produto.custo_unitario)

        total_pedidos += faturamento
        lucro_pedidos += (faturamento - custo)

    total_despesas = 0
    for despesa in despesas:
        total_despesas += float(despesa.valor)

    total_retiradas = 0
    for retirada in retiradas:
        total_retiradas += float(retirada.valor)

    lucro_bruto = lucro_vendas + lucro_pedidos
    lucro_liquido = lucro_bruto - total_despesas - total_retiradas

    return render(request, "lucro_diario.html", {
        "loja": loja,
        "data": hoje,
        "total_vendas": total_vendas,
        "total_pedidos": total_pedidos,
        "total_despesas": total_despesas,
        "total_retiradas": total_retiradas,
        "lucro_vendas": lucro_vendas,
        "lucro_pedidos": lucro_pedidos,
        "lucro_bruto": lucro_bruto,
        "lucro_liquido": lucro_liquido,
        "quantidade_vendas": vendas.count(),
        "quantidade_pedidos": pedidos.count(),
        "quantidade_despesas": despesas.count(),
        "quantidade_retiradas": retiradas.count(),
    })

@login_required
def compras_estoque(request):
    loja = obter_loja_usuario(request.user)

    if not loja:
        return render(request, "erro_loja.html")

    if not usuario_eh_gerente_ou_admin(request.user):
        return render(request, "operacao_bloqueada.html", {
            "mensagem": "Apenas gerente ou admin podem registrar compras de estoque."
        })

    if dia_fechado(loja):
        return render(request, "operacao_bloqueada.html", {
            "mensagem": "O dia de hoje já foi fechado. Não é possível registrar compras de estoque."
        })

    produtos = Produto.objects.filter(loja=loja).order_by("nome")
    fornecedores = Fornecedor.objects.filter(loja=loja, ativo=True).order_by("nome")

    status = request.GET.get("status", "").strip()
    fornecedor_filtro = request.GET.get("fornecedor", "").strip()

    compras = CompraEstoque.objects.filter(loja=loja)

    if status:
        compras = compras.filter(status=status)

    if fornecedor_filtro:
        compras = compras.filter(fornecedor__icontains=fornecedor_filtro)

    compras = compras.order_by("-data")[:100]

    if request.method == "POST":
        produto_id = request.POST.get("produto")
        fornecedor_post = request.POST.get("fornecedor")
        quantidade = request.POST.get("quantidade")
        tipo_compra = request.POST.get("tipo_compra", "troca")
        observacoes = request.POST.get("observacoes")

        try:
            quantidade = int(quantidade)
        except (TypeError, ValueError):
            return render(request, "compras_estoque.html", {
                "loja": loja,
                "produtos": produtos,
                "compras": compras,
                "erro": "Quantidade inválida.",
                "status": status,
                "fornecedor": fornecedor,
            })

        if quantidade <= 0:
            return render(request, "compras_estoque.html", {
                "loja": loja,
                "produtos": produtos,
                "compras": compras,
                "erro": "A quantidade deve ser maior que zero.",
                "status": status,
                "fornecedor": fornecedor,
            })

        produto = Produto.objects.get(id=produto_id, loja=loja)

        if produto.controla_retorno and tipo_compra == "troca" and produto.estoque_vazio < quantidade:
            return render(request, "compras_estoque.html", {
                "loja": loja,
                "produtos": produtos,
                "compras": compras,
                "erro": "Estoque vazio insuficiente para fazer a troca com a distribuidora.",
                "status": status,
                "fornecedor": fornecedor,
            })

        compra = CompraEstoque.objects.create(
            loja=loja,
            produto=produto,
            fornecedor=fornecedor_post,
            quantidade=quantidade,
            tipo_compra=tipo_compra,
            registrado_por=request.user,
            observacoes=observacoes,
            status="pendente",
        )

        # Estoque sobe imediatamente ao registrar a chegada
        produto.estoque_cheio += quantidade
        if produto.controla_retorno and tipo_compra == "troca":
            produto.estoque_vazio -= quantidade
        produto.save()

        registrar_auditoria(
            loja=loja,
            usuario=request.user,
            acao="Compra de estoque registrada",
            descricao=f"Compra #{compra.id} | Produto: {produto.nome} | Quantidade: {quantidade} | Tipo compra: {tipo_compra} | Fornecedor: {fornecedor_post} | Estoque atualizado imediatamente."
        )

        return redirect("/compras-estoque/")

    return render(request, "compras_estoque.html", {
        "loja": loja,
        "produtos": produtos,
        "fornecedores": fornecedores,
        "compras": compras,
        "status": status,
        "fornecedor": fornecedor_filtro,
    })

@login_required
def contas_pagar(request):
    loja = obter_loja_usuario(request.user)

    if not loja:
        return render(request, "erro_loja.html")

    if not usuario_eh_admin(request.user):
        return render(request, "operacao_bloqueada.html", {
            "mensagem": "Apenas administradores podem acessar contas a pagar."
        })

    if request.method == "POST":
        descricao = request.POST.get("descricao")
        categoria = request.POST.get("categoria")
        valor = request.POST.get("valor")
        vencimento = request.POST.get("vencimento")

        ContaPagar.objects.create(
            loja=loja,
            descricao=descricao,
            categoria=categoria,
            valor=valor,
            vencimento=vencimento,
            registrado_por=request.user,
        )

        return redirect("/contas-pagar/")

    contas = ContaPagar.objects.filter(loja=loja).order_by("status", "vencimento")

    total_pendente = 0
    total_pago = 0

    for conta in contas:
        if conta.status == "pendente":
            total_pendente += float(conta.valor)
        else:
            total_pago += float(conta.valor)

    return render(request, "contas_pagar.html", {
        "loja": loja,
        "contas": contas,
        "total_pendente": total_pendente,
        "total_pago": total_pago,
    })


@login_required
def pagar_conta(request, conta_id):
    loja = obter_loja_usuario(request.user)

    if not loja:
        return render(request, "erro_loja.html")

    if not usuario_eh_admin(request.user):
        return render(request, "operacao_bloqueada.html", {
            "mensagem": "Apenas administradores podem alterar contas a pagar."
        })

    conta = ContaPagar.objects.get(id=conta_id, loja=loja)
    conta.status = "pago"
    conta.save()

    return redirect("/contas-pagar/")

@login_required
def dre_mensal(request):
    if not usuario_eh_admin(request.user):
        return render(request, "operacao_bloqueada.html", {
            "mensagem": "Apenas administradores podem acessar o DRE mensal."
        })

    loja = obter_loja_usuario(request.user)

    if not loja:
        return render(request, "erro_loja.html")

    mes = request.GET.get("mes")
    ano = request.GET.get("ano")

    hoje = timezone.now()

    if mes and ano:
        mes = int(mes)
        ano = int(ano)
    else:
        mes = hoje.month
        ano = hoje.year

    vendas = Venda.objects.filter(
        loja=loja,
        status="ativa",
        data_venda__month=mes,
        data_venda__year=ano
    )

    total_vendas = sum(v.total_venda() for v in vendas)

    despesas = Despesa.objects.filter(
        loja=loja,
        data__month=mes,
        data__year=ano
    )

    total_despesas = sum(d.valor for d in despesas)

    retiradas = RetiradaFuncionario.objects.filter(
        loja=loja,
        data__month=mes,
        data__year=ano
    )

    total_retiradas = sum(r.valor for r in retiradas)

    custo_total = sum(
        v.quantidade * v.produto.custo_unitario
        for v in vendas
    )

    lucro_bruto = total_vendas - custo_total
    lucro_liquido = lucro_bruto - total_despesas - total_retiradas

    return render(request, "dre_mensal.html", {
        "loja": loja,
        "mes": mes,
        "ano": ano,
        "total_vendas": total_vendas,
        "custo_total": custo_total,
        "lucro_bruto": lucro_bruto,
        "total_despesas": total_despesas,
        "total_retiradas": total_retiradas,
        "lucro_liquido": lucro_liquido,
    })

@login_required
def salvar_fechamento_caixa(request):
    loja = obter_loja_usuario(request.user)

    if not loja:
        return render(request, "erro_loja.html")

    hoje = timezone.now().date()

    if FechamentoCaixa.objects.filter(loja=loja, data=hoje).exists():
        return render(request, "operacao_bloqueada.html", {
            "mensagem": "O fechamento de hoje já foi salvo anteriormente."
        })

    vendas = Venda.objects.filter(
        data_venda__date=hoje,
        loja=loja,
        status="ativa"
    )

    pedidos = Pedido.objects.filter(
        data_pedido__date=hoje,
        loja=loja
    ).exclude(status="cancelado")

    despesas = Despesa.objects.filter(loja=loja, data__date=hoje)
    retiradas = RetiradaFuncionario.objects.filter(loja=loja, data__date=hoje)

    total_vendas = 0
    for venda in vendas:
        total_vendas += float(venda.quantidade * venda.preco_unitario)

    total_pedidos = 0
    for pedido in pedidos:
        total_pedidos += float(pedido.quantidade * pedido.preco_unitario)

    total_despesas = 0
    for despesa in despesas:
        total_despesas += float(despesa.valor)

    total_retiradas = 0
    for retirada in retiradas:
        total_retiradas += float(retirada.valor)

    dinheiro_vendas, pix_vendas, credito_vendas, debito_vendas = somar_pagamentos_mistos(vendas)
    dinheiro_pedidos, pix_pedidos, credito_pedidos, debito_pedidos = somar_pagamentos_pedidos(pedidos)

    total_geral = total_vendas + total_pedidos - total_despesas - total_retiradas

    FechamentoCaixa.objects.update_or_create(
        loja=loja,
        data=hoje,
        defaults={
            "total_vendas": total_vendas,
            "total_pedidos": total_pedidos,
            "total_despesas": total_despesas,
            "total_retiradas": total_retiradas,
            "total_geral": total_geral,
            "dinheiro_vendas": dinheiro_vendas,
            "pix_vendas": pix_vendas,
            "credito_vendas": credito_vendas,
            "debito_vendas": debito_vendas,
            "dinheiro_pedidos": dinheiro_pedidos,
            "pix_pedidos": pix_pedidos,
            "credito_pedidos": credito_pedidos,
            "debito_pedidos": debito_pedidos,
            "criado_por": request.user,
        }
    )

    registrar_auditoria(
    loja=loja,
    usuario=request.user,
    acao="Fechamento de caixa salvo",
    descricao=f"Data: {hoje.strftime('%d/%m/%Y')} | Total geral: R$ {total_geral}"
    )

    return redirect("/historico-fechamentos/")

@login_required
def historico_fechamentos(request):
    loja = obter_loja_usuario(request.user)

    if not loja:
        return render(request, "erro_loja.html")

    if not usuario_eh_gerente_ou_admin(request.user):
        return render(request, "operacao_bloqueada.html", {
            "mensagem": "Apenas gerente ou admin podem acessar o histórico de fechamentos."
        })

    fechamentos = FechamentoCaixa.objects.filter(loja=loja)

    return render(request, "historico_fechamentos.html", {
        "loja": loja,
        "fechamentos": fechamentos,
    })

@login_required
def saldo_caixa(request):
    loja = obter_loja_usuario(request.user)

    if not loja:
        return render(request, "erro_loja.html")

    if not usuario_eh_admin(request.user):
        return render(request, "operacao_bloqueada.html", {
            "mensagem": "Apenas administradores podem acessar o saldo de caixa."
        })

    fechamentos = FechamentoCaixa.objects.filter(loja=loja).order_by("data")

    saldo_acumulado = 0
    historico = []

    for fechamento in fechamentos:
        saldo_acumulado += float(fechamento.total_geral)

        historico.append({
            "data": fechamento.data,
            "total_vendas": fechamento.total_vendas,
            "total_pedidos": fechamento.total_pedidos,
            "total_despesas": fechamento.total_despesas,
            "total_retiradas": fechamento.total_retiradas,
            "total_geral": fechamento.total_geral,
            "saldo_acumulado": saldo_acumulado,
        })

    return render(request, "saldo_caixa.html", {
        "loja": loja,
        "historico": historico,
        "saldo_atual": saldo_acumulado,
    })

@login_required
def reabrir_fechamento(request, fechamento_id):
    loja = obter_loja_usuario(request.user)

    if not loja:
        return render(request, "erro_loja.html")

    if not usuario_eh_admin(request.user):
        return render(request, "operacao_bloqueada.html", {
            "mensagem": "Apenas administradores podem reabrir fechamentos."
        })

    fechamento = FechamentoCaixa.objects.get(id=fechamento_id, loja=loja)

    if request.method != "POST":
        return render(request, "confirmar_acao.html", {
            "titulo": "Reabrir fechamento",
            "mensagem": "Tem certeza que deseja reabrir este fechamento? Essa ação remove o fechamento salvo.",
            "detalhes": f"Fechamento da data {fechamento.data.strftime('%d/%m/%Y')} | Total geral: R$ {fechamento.total_geral}",
            "voltar_url": "/historico-fechamentos/",
        })

    registrar_auditoria(
        loja=loja,
        usuario=request.user,
        acao="Fechamento reaberto",
        descricao=f"Fechamento da data {fechamento.data.strftime('%d/%m/%Y')} foi reaberto"
    )

    fechamento.delete()
    return redirect("/historico-fechamentos/")

@login_required
def admin_geral(request):
    loja = obter_loja_usuario(request.user)

    if not loja:
        return render(request, "erro_loja.html")

    if not usuario_eh_admin(request.user):
        return render(request, "operacao_bloqueada.html", {
            "mensagem": "Apenas administradores podem acessar o painel administrativo."
        })

    hoje = timezone.now().date()

    vendas = Venda.objects.filter(
        loja=loja,
        data_venda__date=hoje,
        status="ativa"
    ).select_related("produto")

    pedidos = Pedido.objects.filter(
        loja=loja,
        data_pedido__date=hoje
    ).exclude(status="cancelado").select_related("produto")

    despesas = Despesa.objects.filter(
        loja=loja,
        data__date=hoje
    )

    retiradas = RetiradaFuncionario.objects.filter(
        loja=loja,
        data__date=hoje
    )

    contas_pendentes = ContaPagar.objects.filter(
        loja=loja,
        status="pendente"
    ).order_by("vencimento")[:10]

    produtos = Produto.objects.filter(loja=loja)

    estoque_baixo = []
    for produto in produtos:
        if produto.estoque_cheio < produto.alerta_estoque_minimo:
            estoque_baixo.append(produto)

    total_vendas = 0
    lucro_vendas = 0
    for venda in vendas:
        faturamento = float(venda.quantidade * venda.preco_unitario)
        custo = float(venda.quantidade * venda.produto.custo_unitario)
        total_vendas += faturamento
        lucro_vendas += (faturamento - custo)

    total_pedidos = 0
    lucro_pedidos = 0
    for pedido in pedidos:
        faturamento = float(pedido.quantidade * pedido.preco_unitario)
        custo = float(pedido.quantidade * pedido.produto.custo_unitario)
        total_pedidos += faturamento
        lucro_pedidos += (faturamento - custo)

    total_despesas = 0
    for despesa in despesas:
        total_despesas += float(despesa.valor)

    total_retiradas = 0
    for retirada in retiradas:
        total_retiradas += float(retirada.valor)

    lucro_liquido = (lucro_vendas + lucro_pedidos) - total_despesas - total_retiradas

    fechamentos = FechamentoCaixa.objects.filter(loja=loja).order_by("data")
    saldo_acumulado = 0
    for fechamento in fechamentos:
        saldo_acumulado += float(fechamento.total_geral)

    ultimo_fechamento = FechamentoCaixa.objects.filter(loja=loja).order_by("-data", "-criado_em").first()

    total_contas_pendentes = 0
    for conta in ContaPagar.objects.filter(loja=loja, status="pendente"):
        total_contas_pendentes += float(conta.valor)

    compras_pendentes = CompraEstoque.objects.filter(
        loja=loja,
        status="pendente"
    ).order_by("-data")[:10]

    total_compras_pendentes = CompraEstoque.objects.filter(
        loja=loja,
        status="pendente"
    ).count()

    return render(request, "admin_geral.html", {
        "loja": loja,
        "data": hoje,
        "total_vendas": total_vendas,
        "total_pedidos": total_pedidos,
        "total_despesas": total_despesas,
        "total_retiradas": total_retiradas,
        "lucro_liquido": lucro_liquido,
        "saldo_acumulado": saldo_acumulado,
        "total_contas_pendentes": total_contas_pendentes,
        "contas_pendentes": contas_pendentes,
        "estoque_baixo": estoque_baixo,
        "ultimo_fechamento": ultimo_fechamento,
        "quantidade_vendas": vendas.count(),
        "quantidade_pedidos": pedidos.count(),
        "quantidade_despesas": despesas.count(),
        "quantidade_retiradas": retiradas.count(),
        "compras_pendentes": compras_pendentes,
        "total_compras_pendentes": total_compras_pendentes,
    })

@login_required
def comparativo_mensal(request):
    loja = obter_loja_usuario(request.user)

    if not loja:
        return render(request, "erro_loja.html")

    if not usuario_eh_admin(request.user):
        return render(request, "operacao_bloqueada.html", {
            "mensagem": "Apenas administradores podem acessar o comparativo mensal."
        })

    hoje = timezone.now().date()

    mes_atual = hoje.month
    ano_atual = hoje.year

    if mes_atual == 1:
        mes_anterior = 12
        ano_anterior = ano_atual - 1
    else:
        mes_anterior = mes_atual - 1
        ano_anterior = ano_atual

    vendas_atual = Venda.objects.filter(
        loja=loja,
        data_venda__year=ano_atual,
        data_venda__month=mes_atual,
        status="ativa"
    ).select_related("produto")

    pedidos_atual = Pedido.objects.filter(
        loja=loja,
        data_pedido__year=ano_atual,
        data_pedido__month=mes_atual
    ).exclude(status="cancelado").select_related("produto")

    despesas_atual = Despesa.objects.filter(
        loja=loja,
        data__year=ano_atual,
        data__month=mes_atual
    )

    retiradas_atual = RetiradaFuncionario.objects.filter(
        loja=loja,
        data__year=ano_atual,
        data__month=mes_atual
    )

    faturamento_vendas_atual = 0
    lucro_vendas_atual = 0
    for venda in vendas_atual:
        faturamento = float(venda.quantidade * venda.preco_unitario)
        custo = float(venda.quantidade * venda.produto.custo_unitario)
        faturamento_vendas_atual += faturamento
        lucro_vendas_atual += (faturamento - custo)

    faturamento_pedidos_atual = 0
    lucro_pedidos_atual = 0
    for pedido in pedidos_atual:
        faturamento = float(pedido.quantidade * pedido.preco_unitario)
        custo = float(pedido.quantidade * pedido.produto.custo_unitario)
        faturamento_pedidos_atual += faturamento
        lucro_pedidos_atual += (faturamento - custo)

    total_despesas_atual = 0
    for despesa in despesas_atual:
        total_despesas_atual += float(despesa.valor)

    total_retiradas_atual = 0
    for retirada in retiradas_atual:
        total_retiradas_atual += float(retirada.valor)

    lucro_liquido_atual = (
        lucro_vendas_atual + lucro_pedidos_atual
        - total_despesas_atual
        - total_retiradas_atual
    )

    faturamento_total_atual = faturamento_vendas_atual + faturamento_pedidos_atual

    vendas_anterior = Venda.objects.filter(
        loja=loja,
        data_venda__year=ano_anterior,
        data_venda__month=mes_anterior,
        status="ativa"
    ).select_related("produto")

    pedidos_anterior = Pedido.objects.filter(
        loja=loja,
        data_pedido__year=ano_anterior,
        data_pedido__month=mes_anterior
    ).exclude(status="cancelado").select_related("produto")

    despesas_anterior = Despesa.objects.filter(
        loja=loja,
        data__year=ano_anterior,
        data__month=mes_anterior
    )

    retiradas_anterior = RetiradaFuncionario.objects.filter(
        loja=loja,
        data__year=ano_anterior,
        data__month=mes_anterior
    )

    faturamento_vendas_anterior = 0
    lucro_vendas_anterior = 0
    for venda in vendas_anterior:
        faturamento = float(venda.quantidade * venda.preco_unitario)
        custo = float(venda.quantidade * venda.produto.custo_unitario)
        faturamento_vendas_anterior += faturamento
        lucro_vendas_anterior += (faturamento - custo)

    faturamento_pedidos_anterior = 0
    lucro_pedidos_anterior = 0
    for pedido in pedidos_anterior:
        faturamento = float(pedido.quantidade * pedido.preco_unitario)
        custo = float(pedido.quantidade * pedido.produto.custo_unitario)
        faturamento_pedidos_anterior += faturamento
        lucro_pedidos_anterior += (faturamento - custo)

    total_despesas_anterior = 0
    for despesa in despesas_anterior:
        total_despesas_anterior += float(despesa.valor)

    total_retiradas_anterior = 0
    for retirada in retiradas_anterior:
        total_retiradas_anterior += float(retirada.valor)

    lucro_liquido_anterior = (
        lucro_vendas_anterior + lucro_pedidos_anterior
        - total_despesas_anterior
        - total_retiradas_anterior
    )

    faturamento_total_anterior = faturamento_vendas_anterior + faturamento_pedidos_anterior

    def calcular_variacao(atual, anterior):
        diferenca = atual - anterior
        if anterior == 0:
            porcentagem = 100 if atual > 0 else 0
        else:
            porcentagem = (diferenca / anterior) * 100
        return diferenca, porcentagem

    dif_faturamento, perc_faturamento = calcular_variacao(
        faturamento_total_atual, faturamento_total_anterior
    )
    dif_despesas, perc_despesas = calcular_variacao(
        total_despesas_atual, total_despesas_anterior
    )
    dif_retiradas, perc_retiradas = calcular_variacao(
        total_retiradas_atual, total_retiradas_anterior
    )
    dif_lucro, perc_lucro = calcular_variacao(
        lucro_liquido_atual, lucro_liquido_anterior
    )

    nomes_meses = {
        1: "Janeiro", 2: "Fevereiro", 3: "Março", 4: "Abril",
        5: "Maio", 6: "Junho", 7: "Julho", 8: "Agosto",
        9: "Setembro", 10: "Outubro", 11: "Novembro", 12: "Dezembro"
    }

    return render(request, "comparativo_mensal.html", {
        "loja": loja,
        "mes_atual_nome": nomes_meses[mes_atual],
        "ano_atual": ano_atual,
        "mes_anterior_nome": nomes_meses[mes_anterior],
        "ano_anterior": ano_anterior,

        "faturamento_total_atual": faturamento_total_atual,
        "total_despesas_atual": total_despesas_atual,
        "total_retiradas_atual": total_retiradas_atual,
        "lucro_liquido_atual": lucro_liquido_atual,

        "faturamento_total_anterior": faturamento_total_anterior,
        "total_despesas_anterior": total_despesas_anterior,
        "total_retiradas_anterior": total_retiradas_anterior,
        "lucro_liquido_anterior": lucro_liquido_anterior,

        "dif_faturamento": dif_faturamento,
        "perc_faturamento": perc_faturamento,
        "dif_despesas": dif_despesas,
        "perc_despesas": perc_despesas,
        "dif_retiradas": dif_retiradas,
        "perc_retiradas": perc_retiradas,
        "dif_lucro": dif_lucro,
        "perc_lucro": perc_lucro,
    })

@login_required
def aprovar_compra_estoque(request, compra_id):
    loja = obter_loja_usuario(request.user)

    if not loja:
        return render(request, "erro_loja.html")

    if not usuario_eh_admin(request.user):
        return render(request, "operacao_bloqueada.html", {
            "mensagem": "Apenas administradores podem aprovar compras de estoque."
        })

    if dia_fechado(loja):
        return render(request, "operacao_bloqueada.html", {
            "mensagem": "O dia de hoje já foi fechado. Não é possível aprovar compras de estoque."
        })

    compra = CompraEstoque.objects.get(id=compra_id, loja=loja)

    if compra.status == "aprovada":
        return render(request, "operacao_bloqueada.html", {
            "mensagem": "Essa compra já foi aprovada anteriormente."
        })

    if request.method == "GET":
        return render(request, "aprovar_compra_estoque.html", {
            "loja": loja,
            "compra": compra,
        })

    custo_unitario_compra = request.POST.get("custo_unitario_compra")

    try:
        custo_unitario_compra = Decimal(custo_unitario_compra)
    except (InvalidOperation, TypeError):
        return render(request, "aprovar_compra_estoque.html", {
            "loja": loja,
            "compra": compra,
            "erro": "Custo unitário inválido."
        })

    if custo_unitario_compra <= 0:
        return render(request, "aprovar_compra_estoque.html", {
            "loja": loja,
            "compra": compra,
            "erro": "O custo unitário deve ser maior que zero."
        })

    produto = compra.produto
    quantidade_nova = compra.quantidade

    estoque_atual = produto.estoque_cheio
    custo_atual = Decimal(produto.custo_unitario or 0)
    quantidade_total = estoque_atual + quantidade_nova

    if quantidade_total > 0:
        novo_custo_medio = (
            (Decimal(estoque_atual) * custo_atual) +
            (Decimal(quantidade_nova) * custo_unitario_compra)
        ) / Decimal(quantidade_total)
    else:
        novo_custo_medio = custo_unitario_compra

    custo_total = Decimal(quantidade_nova) * custo_unitario_compra

    compra.custo_unitario_compra = custo_unitario_compra
    compra.custo_total = custo_total
    compra.status = "aprovada"
    compra.aprovado_por = request.user
    compra.aprovado_em = timezone.now()
    compra.save()

    # Estoque já foi atualizado no momento do registro.
    # Aqui apenas atualizamos o custo médio do produto.
    produto.custo_unitario = novo_custo_medio.quantize(Decimal("0.01"))
    produto.save()

    registrar_auditoria(
        loja=loja,
        usuario=request.user,
        acao="Compra de estoque aprovada",
        descricao=f"Compra #{compra.id} | Produto: {produto.nome} | Quantidade: {quantidade_nova} | Tipo compra: {compra.tipo_compra} | Custo unitário: R$ {custo_unitario_compra}"
    )

    return redirect("/compras-estoque/")


@login_required
def reprovar_compra_estoque(request, compra_id):
    loja = obter_loja_usuario(request.user)

    if not loja:
        return render(request, "erro_loja.html")

    if not usuario_eh_admin(request.user):
        return render(request, "operacao_bloqueada.html", {
            "mensagem": "Apenas administradores podem reprovar compras de estoque."
        })

    if request.method != "POST":
        return redirect("/compras-estoque/")

    try:
        compra = CompraEstoque.objects.get(id=compra_id, loja=loja)
    except CompraEstoque.DoesNotExist:
        return render(request, "operacao_bloqueada.html", {
            "mensagem": "Compra não encontrada."
        })

    if compra.status != "pendente":
        return render(request, "operacao_bloqueada.html", {
            "mensagem": "Só é possível reprovar compras com status Pendente."
        })

    produto = compra.produto
    quantidade = compra.quantidade

    # Reverter estoque que foi aplicado no registro
    produto.estoque_cheio -= quantidade
    if produto.controla_retorno and compra.tipo_compra == "troca":
        produto.estoque_vazio += quantidade
    produto.save()

    compra.status = "reprovada"
    compra.aprovado_por = request.user
    compra.aprovado_em = timezone.now()
    compra.save()

    registrar_auditoria(
        loja=loja,
        usuario=request.user,
        acao="Compra de estoque reprovada",
        descricao=f"Compra #{compra.id} | Produto: {produto.nome} | Quantidade: {quantidade} | Estoque revertido."
    )

    return redirect("/compras-estoque/")


@login_required
def estoque_admin(request):
    loja = obter_loja_usuario(request.user)

    if not loja:
        return render(request, "erro_loja.html")

    if not usuario_eh_gerente_ou_admin(request.user):
        return render(request, "operacao_bloqueada.html", {
            "mensagem": "Apenas gerente ou admin podem acessar o painel de estoque."
        })

    produtos = Produto.objects.filter(loja=loja).order_by("nome")

    lista_estoque = []

    for produto in produtos:
        compras_pendentes = CompraEstoque.objects.filter(
            loja=loja,
            produto=produto,
            status="pendente"
        ).count()

        if produto.controla_retorno:
            if produto.estoque_cheio <= 0:
                status_estoque = "Sem cheio"
            elif produto.estoque_cheio < produto.alerta_estoque_minimo:
                status_estoque = "Baixo"
            else:
                status_estoque = "OK"
        else:
            if produto.estoque_cheio <= 0:
                status_estoque = "Sem estoque"
            elif produto.estoque_cheio < produto.alerta_estoque_minimo:
                status_estoque = "Baixo"
            else:
                status_estoque = "OK"

        lista_estoque.append({
            "produto": produto,
            "tipo_produto": "Retornável" if produto.controla_retorno else "Simples",
            "estoque_cheio": produto.estoque_cheio,
            "estoque_vazio": produto.estoque_vazio,
            "estoque_total": produto.estoque_cheio + produto.estoque_vazio,
            "alerta_minimo": produto.alerta_estoque_minimo,
            "custo_unitario": produto.custo_unitario if usuario_eh_admin(request.user) else None,
            "status_estoque": status_estoque,
            "compras_pendentes": compras_pendentes,
        })

    total_cheio = sum(item["estoque_cheio"] for item in lista_estoque)
    total_vazio = sum(item["estoque_vazio"] for item in lista_estoque)
    total_pendencias = sum(item["compras_pendentes"] for item in lista_estoque)

    produtos_retornaveis = sum(1 for item in lista_estoque if item["produto"].controla_retorno)
    produtos_simples = sum(1 for item in lista_estoque if not item["produto"].controla_retorno)

    return render(request, "estoque_admin.html", {
        "loja": loja,
        "lista_estoque": lista_estoque,
        "total_cheio": total_cheio,
        "total_vazio": total_vazio,
        "total_pendencias": total_pendencias,
        "produtos_retornaveis": produtos_retornaveis,
        "produtos_simples": produtos_simples,
    })

@login_required
def inicio(request):
    loja = obter_loja_usuario(request.user)

    if not loja:
        return render(request, "erro_loja.html")

    if usuario_eh_admin(request.user):
        return redirect("/admin-geral/")

    if usuario_eh_gerente(request.user):
        return redirect("/dashboard/")

    return render(request, "home_funcionario.html", {
        "loja": loja,
    })

@login_required
def relatorio_vendas(request):
    loja = obter_loja_usuario(request.user)

    if not loja:
        return render(request, "erro_loja.html")

    vendas = Venda.objects.filter(
        loja=loja
    ).order_by("-data_venda")[:100]

    total = 0
    for venda in vendas:
        if venda.status == "ativa":
            total += float(venda.quantidade * venda.preco_unitario)

    return render(request, "relatorio.html", {
        "loja": loja,
        "vendas": vendas,
        "total": total,
    })

@login_required
def auditoria_sistema(request):
    loja = obter_loja_usuario(request.user)

    if not loja:
        return render(request, "erro_loja.html")

    if not usuario_eh_admin(request.user):
        return render(request, "operacao_bloqueada.html", {
            "mensagem": "Apenas administradores podem acessar a auditoria do sistema."
        })

    auditorias = AuditoriaSistema.objects.filter(loja=loja).order_by("-data")[:200]

    return render(request, "auditoria_sistema.html", {
        "loja": loja,
        "auditorias": auditorias,
    })

@login_required
def inventario_estoque(request):
    loja = obter_loja_usuario(request.user)

    if not loja:
        return render(request, "erro_loja.html")

    if not usuario_eh_admin(request.user):
        return render(request, "operacao_bloqueada.html", {
            "mensagem": "Apenas administradores podem fazer inventário de estoque."
        })

    produtos = Produto.objects.filter(loja=loja).order_by("nome")

    if request.method == "POST":
        observacoes = request.POST.get("observacoes", "").strip()

        inventario = InventarioEstoque.objects.create(
            loja=loja,
            criado_por=request.user,
            observacoes=observacoes,
        )

        for produto in produtos:
            cheio_contado = request.POST.get(f"cheio_{produto.id}", "0")
            vazio_contado = request.POST.get(f"vazio_{produto.id}", "0")

            try:
                cheio_contado = int(cheio_contado or 0)
                vazio_contado = int(vazio_contado or 0)
            except ValueError:
                cheio_contado = 0
                vazio_contado = 0

            cheio_sistema = produto.estoque_cheio
            vazio_sistema = produto.estoque_vazio

            diferenca_cheio = cheio_contado - cheio_sistema
            diferenca_vazio = vazio_contado - vazio_sistema

            ItemInventarioEstoque.objects.create(
                inventario=inventario,
                produto=produto,
                estoque_cheio_sistema=cheio_sistema,
                estoque_vazio_sistema=vazio_sistema,
                estoque_cheio_contado=cheio_contado,
                estoque_vazio_contado=vazio_contado,
                diferenca_cheio=diferenca_cheio,
                diferenca_vazio=diferenca_vazio,
            )

            produto.estoque_cheio = cheio_contado
            produto.estoque_vazio = vazio_contado
            produto.save()

            if diferenca_cheio != 0:
                MovimentacaoEstoque.objects.create(
                    loja=loja,
                    produto=produto,
                    usuario=request.user,
                    tipo="ajuste_cheio",
                    quantidade=cheio_contado,
                    motivo=f"Inventário #{inventario.id} | Ajuste de cheio | Sistema: {cheio_sistema} | Contado: {cheio_contado}"
                )

            if diferenca_vazio != 0:
                MovimentacaoEstoque.objects.create(
                    loja=loja,
                    produto=produto,
                    usuario=request.user,
                    tipo="ajuste_vazio",
                    quantidade=vazio_contado,
                    motivo=f"Inventário #{inventario.id} | Ajuste de vazio | Sistema: {vazio_sistema} | Contado: {vazio_contado}"
                )

        registrar_auditoria(
            loja=loja,
            usuario=request.user,
            acao="Inventário de estoque realizado",
            descricao=f"Inventário #{inventario.id} realizado com {produtos.count()} produtos"
        )

        return redirect(f"/inventario-estoque/{inventario.id}/")

    return render(request, "inventario_estoque.html", {
        "loja": loja,
        "produtos": produtos,
    })

@login_required
def detalhe_inventario_estoque(request, inventario_id):
    loja = obter_loja_usuario(request.user)

    if not loja:
        return render(request, "erro_loja.html")

    if not usuario_eh_admin(request.user):
        return render(request, "operacao_bloqueada.html", {
            "mensagem": "Apenas administradores podem acessar inventários."
        })

    inventario = InventarioEstoque.objects.get(id=inventario_id, loja=loja)
    itens = inventario.itens.select_related("produto").all()

    return render(request, "detalhe_inventario_estoque.html", {
        "loja": loja,
        "inventario": inventario,
        "itens": itens,
    })

@login_required
def historico_produto(request, produto_id):
    loja = obter_loja_usuario(request.user)

    if not loja:
        return render(request, "erro_loja.html")

    if not usuario_eh_gerente_ou_admin(request.user):
        return render(request, "operacao_bloqueada.html", {
            "mensagem": "Apenas gerente ou admin podem acessar o histórico do produto."
        })

    try:
        produto = Produto.objects.get(id=produto_id, loja=loja)
    except Produto.DoesNotExist:
        return render(request, "operacao_bloqueada.html", {
            "mensagem": "Produto não encontrado."
        })

    vendas = Venda.objects.filter(
        loja=loja,
        produto=produto
    ).order_by("-data_venda")[:50]

    compras = CompraEstoque.objects.filter(
        loja=loja,
        produto=produto
    ).order_by("-data")[:50]

    movimentacoes = MovimentacaoEstoque.objects.filter(
        loja=loja,
        produto=produto
    ).order_by("-data_movimentacao")[:100]

    inventarios = ItemInventarioEstoque.objects.select_related("inventario").filter(
    produto=produto,
    inventario__loja=loja
    ).order_by("-inventario__data")[:30]

    return render(request, "historico_produto.html", {
        "loja": loja,
        "produto": produto,
        "vendas": vendas,
        "compras": compras,
        "movimentacoes": movimentacoes,
        "inventarios": inventarios,
    })

@login_required
def historico_produto_excel(request, produto_id):
    loja = obter_loja_usuario(request.user)

    if not loja:
        return render(request, "erro_loja.html")

    try:
        produto = Produto.objects.get(id=produto_id, loja=loja)
    except Produto.DoesNotExist:
        return render(request, "operacao_bloqueada.html", {
            "mensagem": "Produto não encontrado."
        })

    movimentacoes = MovimentacaoEstoque.objects.filter(
        loja=loja,
        produto=produto
    ).order_by("-data_movimentacao")

    wb = Workbook()
    ws = wb.active
    ws.title = "Historico Produto"

    negrito = Font(bold=True)

    ws.append(["Data", "Tipo", "Quantidade", "Usuário", "Motivo"])
    for cell in ws[1]:
        cell.font = negrito

    for mov in movimentacoes:
        ws.append([
            mov.data_movimentacao.strftime("%d/%m/%Y %H:%M") if mov.data_movimentacao else "",
            mov.get_tipo_display() if hasattr(mov, "get_tipo_display") else mov.tipo,
            mov.quantidade,
            mov.usuario.username if getattr(mov, "usuario", None) else "",
            mov.motivo or "",
        ])

    for coluna in ws.columns:
        maior = 0
        letra = coluna[0].column_letter
        for celula in coluna:
            valor = str(celula.value) if celula.value is not None else ""
            if len(valor) > maior:
                maior = len(valor)
        ws.column_dimensions[letra].width = min(maior + 2, 40)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="historico_produto_{produto.id}.xlsx"'

    wb.save(response)
    return response


# ===================== VALE GÁS =====================

def saldo_vale_gas(loja, cliente):
    creditos = ValeGas.objects.filter(loja=loja, cliente=cliente, tipo="credito").aggregate(
        total=models.Sum("valor")
    )["total"] or Decimal("0.00")
    debitos = ValeGas.objects.filter(loja=loja, cliente=cliente, tipo="debito").aggregate(
        total=models.Sum("valor")
    )["total"] or Decimal("0.00")
    return creditos - debitos


@login_required
def vale_gas(request):
    loja = obter_loja_usuario(request.user)

    if not loja:
        return render(request, "erro_loja.html")

    if not usuario_eh_gerente_ou_admin(request.user):
        return render(request, "operacao_bloqueada.html", {
            "mensagem": "Apenas gerente ou admin podem gerenciar vale gás."
        })

    clientes = Cliente.objects.filter(loja=loja, ativo=True).order_by("nome")
    erro = ""

    if request.method == "POST":
        cliente_id = request.POST.get("cliente")
        tipo = request.POST.get("tipo")
        valor = request.POST.get("valor", "").strip()
        descricao = request.POST.get("descricao", "").strip()

        if not cliente_id or not tipo or not valor:
            erro = "Preencha todos os campos obrigatórios."
        else:
            try:
                valor_decimal = Decimal(valor.replace(",", "."))
                if valor_decimal <= 0:
                    erro = "O valor deve ser maior que zero."
                else:
                    cliente = Cliente.objects.get(id=cliente_id, loja=loja)

                    if tipo == "debito":
                        saldo = saldo_vale_gas(loja, cliente)
                        if valor_decimal > saldo:
                            erro = f"Saldo insuficiente. Saldo atual: R$ {saldo:.2f}"

                    if not erro:
                        ValeGas.objects.create(
                            loja=loja,
                            cliente=cliente,
                            tipo=tipo,
                            valor=valor_decimal,
                            descricao=descricao,
                            registrado_por=request.user,
                        )

                        registrar_auditoria(
                            loja=loja,
                            usuario=request.user,
                            acao=f"Vale gás - {tipo}",
                            descricao=f"Cliente: {cliente.nome} | Valor: R$ {valor_decimal} | {descricao}"
                        )

                        return redirect("/vale-gas/")
            except (InvalidOperation, TypeError, ValueError):
                erro = "Valor inválido."
            except Cliente.DoesNotExist:
                erro = "Cliente não encontrado."

    cliente_filtro = request.GET.get("cliente", "").strip()
    movimentacoes = ValeGas.objects.filter(loja=loja).select_related("cliente", "registrado_por").order_by("-data")[:100]

    if cliente_filtro:
        movimentacoes = movimentacoes.filter(cliente_id=cliente_filtro)

    clientes_com_saldo = []
    for cliente in clientes:
        saldo = saldo_vale_gas(loja, cliente)
        if saldo > 0:
            clientes_com_saldo.append({
                "cliente": cliente,
                "saldo": saldo,
            })

    total_creditos = sum(c["saldo"] for c in clientes_com_saldo)

    return render(request, "vale_gas.html", {
        "loja": loja,
        "clientes": clientes,
        "movimentacoes": movimentacoes,
        "clientes_com_saldo": clientes_com_saldo,
        "total_creditos": total_creditos,
        "erro": erro,
    })


# ===================== COMODATO =====================

@login_required
def comodatos(request):
    loja = obter_loja_usuario(request.user)

    if not loja:
        return render(request, "erro_loja.html")

    if not usuario_eh_gerente_ou_admin(request.user):
        return render(request, "operacao_bloqueada.html", {
            "mensagem": "Apenas gerente ou admin podem gerenciar comodatos."
        })

    clientes = Cliente.objects.filter(loja=loja, ativo=True).order_by("nome")
    erro = ""

    if request.method == "POST":
        acao = request.POST.get("acao", "novo")

        if acao == "devolver":
            comodato_id = request.POST.get("comodato_id")
            try:
                comodato = Comodato.objects.get(id=comodato_id, loja=loja)
                comodato.status = "devolvido"
                comodato.data_devolucao = timezone.now().date()
                comodato.save()

                registrar_auditoria(
                    loja=loja,
                    usuario=request.user,
                    acao="Comodato devolvido",
                    descricao=f"Item: {comodato.item} | Cliente: {comodato.cliente.nome}"
                )
            except Comodato.DoesNotExist:
                erro = "Comodato não encontrado."

            return redirect("/comodatos/")

        elif acao == "perdido":
            comodato_id = request.POST.get("comodato_id")
            try:
                comodato = Comodato.objects.get(id=comodato_id, loja=loja)
                comodato.status = "perdido"
                comodato.save()

                registrar_auditoria(
                    loja=loja,
                    usuario=request.user,
                    acao="Comodato marcado como perdido",
                    descricao=f"Item: {comodato.item} | Cliente: {comodato.cliente.nome}"
                )
            except Comodato.DoesNotExist:
                erro = "Comodato não encontrado."

            return redirect("/comodatos/")

        else:
            cliente_id = request.POST.get("cliente")
            item = request.POST.get("item", "").strip()
            quantidade = request.POST.get("quantidade", "1").strip()
            observacoes = request.POST.get("observacoes", "").strip()

            if not cliente_id or not item:
                erro = "Preencha o cliente e o item."
            else:
                try:
                    quantidade_int = int(quantidade)
                    if quantidade_int <= 0:
                        erro = "A quantidade deve ser maior que zero."
                    else:
                        cliente = Cliente.objects.get(id=cliente_id, loja=loja)

                        Comodato.objects.create(
                            loja=loja,
                            cliente=cliente,
                            item=item,
                            quantidade=quantidade_int,
                            observacoes=observacoes,
                            registrado_por=request.user,
                        )

                        registrar_auditoria(
                            loja=loja,
                            usuario=request.user,
                            acao="Comodato registrado",
                            descricao=f"Item: {item} | Qtd: {quantidade_int} | Cliente: {cliente.nome}"
                        )

                        return redirect("/comodatos/")
                except (ValueError, TypeError):
                    erro = "Quantidade inválida."
                except Cliente.DoesNotExist:
                    erro = "Cliente não encontrado."

    status_filtro = request.GET.get("status", "ativo").strip()
    comodatos_qs = Comodato.objects.filter(loja=loja).select_related("cliente", "registrado_por")

    if status_filtro:
        comodatos_qs = comodatos_qs.filter(status=status_filtro)

    comodatos_qs = comodatos_qs.order_by("-data_saida")

    total_ativos = Comodato.objects.filter(loja=loja, status="ativo").count()
    total_devolvidos = Comodato.objects.filter(loja=loja, status="devolvido").count()
    total_perdidos = Comodato.objects.filter(loja=loja, status="perdido").count()

    return render(request, "comodatos.html", {
        "loja": loja,
        "clientes": clientes,
        "comodatos": comodatos_qs,
        "status_filtro": status_filtro,
        "total_ativos": total_ativos,
        "total_devolvidos": total_devolvidos,
        "total_perdidos": total_perdidos,
        "erro": erro,
    })


@login_required
def comodato_pdf(request, comodato_id):
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas
    from reportlab.lib.units import cm

    loja = obter_loja_usuario(request.user)
    if not loja:
        return render(request, "erro_loja.html")

    comodato = get_object_or_404(Comodato, id=comodato_id, loja=loja)
    cliente = comodato.cliente

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f'inline; filename="comodato_{comodato.id}.pdf"'

    p = canvas.Canvas(response, pagesize=A4)
    largura, altura = A4

    def linha(y, texto, tamanho=11, negrito=False):
        if negrito:
            p.setFont("Helvetica-Bold", tamanho)
        else:
            p.setFont("Helvetica", tamanho)
        p.drawString(2 * cm, y, texto)

    y = altura - 2 * cm

    # Cabeçalho
    p.setFont("Helvetica-Bold", 18)
    p.drawString(2 * cm, y, loja.nome)
    y -= 0.7 * cm
    p.setFont("Helvetica", 10)
    p.drawString(2 * cm, y, loja.endereco or "")
    y -= 1.2 * cm

    # Título
    p.setFont("Helvetica-Bold", 14)
    p.drawCentredString(largura / 2, y, "CONTRATO DE COMODATO DE BOTIJÃO")
    y -= 0.5 * cm
    p.setLineWidth(1.5)
    p.line(2 * cm, y, largura - 2 * cm, y)
    y -= 0.8 * cm

    # Dados do comodato
    linha(y, f"Nº do Contrato: {comodato.id:04d}    |    Data: {comodato.data_saida.strftime('%d/%m/%Y')}", negrito=True)
    y -= 0.7 * cm

    linha(y, "COMODANTE (Empresa):", negrito=True)
    y -= 0.5 * cm
    linha(y, f"  {loja.nome}  —  {loja.endereco or ''}")
    y -= 0.8 * cm

    linha(y, "COMODATÁRIO (Cliente):", negrito=True)
    y -= 0.5 * cm
    linha(y, f"  Nome: {cliente.nome}")
    y -= 0.45 * cm
    linha(y, f"  Telefone: {cliente.telefone or '-'}    CPF: {cliente.cpf_cnpj or '-'}")
    y -= 0.45 * cm
    linha(y, f"  Endereço: {cliente.endereco or '-'}")
    y -= 0.8 * cm

    linha(y, "ITEM CEDIDO EM COMODATO:", negrito=True)
    y -= 0.5 * cm
    item_display = dict(comodato.ITEM_CHOICES).get(comodato.item, comodato.item)
    linha(y, f"  {item_display}    Quantidade: {comodato.quantidade} unidade(s)")
    y -= 0.8 * cm

    # Cláusulas
    linha(y, "CLÁUSULAS:", negrito=True)
    y -= 0.5 * cm

    clausulas = [
        "1. O COMODATÁRIO recebe o(s) botijão(ões) acima identificado(s) em regime de comodato, ou seja,",
        "   a título gratuito e com obrigação de devolvê-lo(s) nas mesmas condições recebidas.",
        "2. O COMODATÁRIO é responsável pela guarda, conservação e uso correto do botijão durante o período",
        "   em que estiver sob sua posse.",
        "3. Em caso de perda, roubo, dano ou deterioração do botijão, o COMODATÁRIO se compromete a",
        "   ressarcir o COMODANTE pelo valor de reposição do bem.",
        "4. O COMODANTE poderá solicitar a devolução do botijão a qualquer momento, mediante aviso prévio.",
        "5. O botijão é de propriedade exclusiva do COMODANTE e não poderá ser alienado, empenhado ou",
        "   cedido pelo COMODATÁRIO a terceiros sem autorização expressa.",
    ]

    for cl in clausulas:
        linha(y, cl, tamanho=10)
        y -= 0.45 * cm

    if comodato.observacoes:
        y -= 0.3 * cm
        linha(y, f"Observações: {comodato.observacoes}", tamanho=10)
        y -= 0.5 * cm

    # Assinaturas
    y -= 1.0 * cm
    p.setLineWidth(0.5)
    p.line(2 * cm, y, 9 * cm, y)
    p.line(11 * cm, y, 18 * cm, y)
    y -= 0.4 * cm
    p.setFont("Helvetica", 9)
    p.drawString(2 * cm, y, "Assinatura do Comodante")
    p.drawString(11 * cm, y, "Assinatura do Comodatário")
    y -= 0.8 * cm
    p.setFont("Helvetica", 9)
    p.drawString(2 * cm, y, f"Data: ___/___/______")
    p.drawString(11 * cm, y, f"Data: ___/___/______")

    p.showPage()
    p.save()
    return response


# ===================== CONTAS A RECEBER =====================

@login_required
def contas_receber(request):
    loja = obter_loja_usuario(request.user)

    if not loja:
        return render(request, "erro_loja.html")

    if not usuario_eh_admin(request.user):
        return render(request, "operacao_bloqueada.html", {
            "mensagem": "Apenas administradores podem acessar contas a receber."
        })

    clientes = Cliente.objects.filter(loja=loja, ativo=True).order_by("nome")

    if request.method == "POST":
        cliente_id = request.POST.get("cliente")
        descricao = request.POST.get("descricao", "").strip()
        valor = request.POST.get("valor", "").strip()
        vencimento = request.POST.get("vencimento", "").strip()

        if not cliente_id or not descricao or not valor or not vencimento:
            return render(request, "contas_receber.html", {
                "loja": loja,
                "clientes": clientes,
                "contas": ContaReceber.objects.filter(loja=loja).order_by("status", "vencimento"),
                "erro": "Preencha todos os campos.",
                "total_pendente": 0,
                "total_recebido": 0,
            })

        try:
            cliente = Cliente.objects.get(id=cliente_id, loja=loja)
            ContaReceber.objects.create(
                loja=loja,
                cliente=cliente,
                descricao=descricao,
                valor=Decimal(valor.replace(",", ".")),
                vencimento=datetime.strptime(vencimento, "%Y-%m-%d").date(),
                registrado_por=request.user,
            )

            registrar_auditoria(
                loja=loja,
                usuario=request.user,
                acao="Conta a receber registrada",
                descricao=f"Cliente: {cliente.nome} | Valor: R$ {valor}"
            )
        except (InvalidOperation, ValueError, Cliente.DoesNotExist):
            pass

        return redirect("/contas-receber/")

    contas = ContaReceber.objects.filter(loja=loja).select_related("cliente").order_by("status", "vencimento")

    total_pendente = sum(float(c.valor) for c in contas if c.status == "pendente")
    total_recebido = sum(float(c.valor) for c in contas if c.status == "recebido")

    return render(request, "contas_receber.html", {
        "loja": loja,
        "clientes": clientes,
        "contas": contas,
        "total_pendente": total_pendente,
        "total_recebido": total_recebido,
    })


@login_required
def receber_conta(request, conta_id):
    loja = obter_loja_usuario(request.user)

    if not loja:
        return render(request, "erro_loja.html")

    if not usuario_eh_admin(request.user):
        return render(request, "operacao_bloqueada.html", {
            "mensagem": "Apenas administradores podem alterar contas a receber."
        })

    conta = ContaReceber.objects.get(id=conta_id, loja=loja)
    conta.status = "recebido"
    conta.recebido_em = timezone.now()
    conta.save()

    registrar_auditoria(
        loja=loja,
        usuario=request.user,
        acao="Conta recebida",
        descricao=f"Cliente: {conta.cliente.nome} | Valor: R$ {conta.valor}"
    )

    return redirect("/contas-receber/")


# ===================== VEÍCULOS =====================

@login_required
def veiculos(request):
    loja = obter_loja_usuario(request.user)

    if not loja:
        return render(request, "erro_loja.html")

    if not usuario_eh_gerente_ou_admin(request.user):
        return render(request, "operacao_bloqueada.html", {
            "mensagem": "Apenas gerente ou admin podem gerenciar veículos."
        })

    entregadores = Entregador.objects.filter(loja=loja, ativo=True).order_by("nome")
    erro = ""

    if request.method == "POST":
        acao = request.POST.get("acao", "novo_veiculo")

        if acao == "novo_veiculo":
            placa = request.POST.get("placa", "").strip().upper()
            modelo = request.POST.get("modelo", "").strip()
            ano = request.POST.get("ano", "").strip()
            cor = request.POST.get("cor", "").strip()
            km_atual = request.POST.get("km_atual", "0").strip()
            motorista_id = request.POST.get("motorista_padrao", "").strip()
            observacoes = request.POST.get("observacoes", "").strip()

            if not placa or not modelo:
                erro = "Placa e modelo são obrigatórios."
            else:
                motorista = None
                if motorista_id:
                    motorista = Entregador.objects.filter(id=motorista_id, loja=loja).first()

                Veiculo.objects.create(
                    loja=loja,
                    placa=placa,
                    modelo=modelo,
                    ano=int(ano) if ano else None,
                    cor=cor,
                    km_atual=int(km_atual) if km_atual else 0,
                    motorista_padrao=motorista,
                    observacoes=observacoes,
                )

                registrar_auditoria(
                    loja=loja,
                    usuario=request.user,
                    acao="Veículo cadastrado",
                    descricao=f"Placa: {placa} | Modelo: {modelo}"
                )
                return redirect("/veiculos/")

        elif acao == "abastecimento":
            veiculo_id = request.POST.get("veiculo")
            km = request.POST.get("km_abastecimento", "").strip()
            litros = request.POST.get("litros", "").strip()
            valor_total = request.POST.get("valor_total", "").strip()
            tipo_combustivel = request.POST.get("tipo_combustivel", "gasolina").strip()

            if not veiculo_id or not km or not litros or not valor_total:
                erro = "Preencha todos os campos do abastecimento."
            else:
                try:
                    veiculo = Veiculo.objects.get(id=veiculo_id, loja=loja)
                    km_int = int(km)

                    AbastecimentoVeiculo.objects.create(
                        loja=loja,
                        veiculo=veiculo,
                        km_abastecimento=km_int,
                        litros=Decimal(litros.replace(",", ".")),
                        valor_total=Decimal(valor_total.replace(",", ".")),
                        tipo_combustivel=tipo_combustivel,
                        registrado_por=request.user,
                    )

                    if km_int > veiculo.km_atual:
                        veiculo.km_atual = km_int
                        veiculo.save()

                    registrar_auditoria(
                        loja=loja,
                        usuario=request.user,
                        acao="Abastecimento registrado",
                        descricao=f"Veículo: {veiculo.placa} | {litros}L | R$ {valor_total}"
                    )
                    return redirect("/veiculos/")
                except (Veiculo.DoesNotExist, ValueError, InvalidOperation):
                    erro = "Dados inválidos no abastecimento."

        elif acao == "manutencao":
            veiculo_id = request.POST.get("veiculo")
            tipo = request.POST.get("tipo_manutencao", "").strip()
            descricao_m = request.POST.get("descricao_manutencao", "").strip()
            valor_m = request.POST.get("valor_manutencao", "").strip()
            km_m = request.POST.get("km_manutencao", "").strip()
            prox_km = request.POST.get("proxima_manutencao_km", "").strip()
            prox_data = request.POST.get("proxima_manutencao_data", "").strip()

            if not veiculo_id or not tipo or not descricao_m or not valor_m:
                erro = "Preencha todos os campos obrigatórios da manutenção."
            else:
                try:
                    veiculo = Veiculo.objects.get(id=veiculo_id, loja=loja)

                    ManutencaoVeiculo.objects.create(
                        loja=loja,
                        veiculo=veiculo,
                        tipo=tipo,
                        descricao=descricao_m,
                        valor=Decimal(valor_m.replace(",", ".")),
                        km_manutencao=int(km_m) if km_m else None,
                        proxima_manutencao_km=int(prox_km) if prox_km else None,
                        proxima_manutencao_data=datetime.strptime(prox_data, "%Y-%m-%d").date() if prox_data else None,
                        registrado_por=request.user,
                    )

                    registrar_auditoria(
                        loja=loja,
                        usuario=request.user,
                        acao="Manutenção registrada",
                        descricao=f"Veículo: {veiculo.placa} | Tipo: {tipo} | R$ {valor_m}"
                    )
                    return redirect("/veiculos/")
                except (Veiculo.DoesNotExist, ValueError, InvalidOperation):
                    erro = "Dados inválidos na manutenção."

    veiculos_lista = Veiculo.objects.filter(loja=loja).select_related("motorista_padrao").order_by("placa")

    alertas_manutencao = []
    for v in veiculos_lista:
        ultima = ManutencaoVeiculo.objects.filter(veiculo=v).order_by("-data").first()
        if ultima and ultima.proxima_manutencao_km and v.km_atual >= ultima.proxima_manutencao_km:
            alertas_manutencao.append({
                "veiculo": v,
                "tipo": ultima.get_tipo_display(),
                "km_previsto": ultima.proxima_manutencao_km,
            })
        if ultima and ultima.proxima_manutencao_data and ultima.proxima_manutencao_data <= timezone.now().date():
            alertas_manutencao.append({
                "veiculo": v,
                "tipo": ultima.get_tipo_display(),
                "data_prevista": ultima.proxima_manutencao_data,
            })

    abastecimentos = AbastecimentoVeiculo.objects.filter(loja=loja).select_related("veiculo").order_by("-data")[:20]
    manutencoes = ManutencaoVeiculo.objects.filter(loja=loja).select_related("veiculo").order_by("-data")[:20]

    total_combustivel_mes = AbastecimentoVeiculo.objects.filter(
        loja=loja,
        data__month=timezone.now().month,
        data__year=timezone.now().year,
    ).aggregate(total=models.Sum("valor_total"))["total"] or Decimal("0.00")

    total_manutencao_mes = ManutencaoVeiculo.objects.filter(
        loja=loja,
        data__month=timezone.now().month,
        data__year=timezone.now().year,
    ).aggregate(total=models.Sum("valor"))["total"] or Decimal("0.00")

    return render(request, "veiculos.html", {
        "loja": loja,
        "veiculos": veiculos_lista,
        "entregadores": entregadores,
        "abastecimentos": abastecimentos,
        "manutencoes": manutencoes,
        "alertas_manutencao": alertas_manutencao,
        "total_combustivel_mes": total_combustivel_mes,
        "total_manutencao_mes": total_manutencao_mes,
        "erro": erro,
    })


@login_required
def editar_veiculo(request, veiculo_id):
    loja = obter_loja_usuario(request.user)

    if not loja:
        return render(request, "erro_loja.html")

    if not usuario_eh_gerente_ou_admin(request.user):
        return render(request, "operacao_bloqueada.html", {
            "mensagem": "Apenas gerente ou admin podem editar veículos."
        })

    veiculo = get_object_or_404(Veiculo, id=veiculo_id, loja=loja)
    entregadores = Entregador.objects.filter(loja=loja, ativo=True).order_by("nome")

    if request.method == "POST":
        veiculo.placa = request.POST.get("placa", "").strip().upper()
        veiculo.modelo = request.POST.get("modelo", "").strip()
        ano = request.POST.get("ano", "").strip()
        veiculo.ano = int(ano) if ano else None
        veiculo.cor = request.POST.get("cor", "").strip()
        km = request.POST.get("km_atual", "0").strip()
        veiculo.km_atual = int(km) if km else 0
        veiculo.status = request.POST.get("status", "ativo")
        veiculo.observacoes = request.POST.get("observacoes", "").strip()

        motorista_id = request.POST.get("motorista_padrao", "").strip()
        if motorista_id:
            veiculo.motorista_padrao = Entregador.objects.filter(id=motorista_id, loja=loja).first()
        else:
            veiculo.motorista_padrao = None

        veiculo.save()

        registrar_auditoria(
            loja=loja,
            usuario=request.user,
            acao="Veículo editado",
            descricao=f"Placa: {veiculo.placa} | Status: {veiculo.status}"
        )
        return redirect("/veiculos/")

    return render(request, "editar_veiculo.html", {
        "loja": loja,
        "veiculo": veiculo,
        "entregadores": entregadores,
    })


# ===================== VENDAS ANTECIPADAS =====================

@login_required
def vendas_antecipadas(request):
    loja = obter_loja_usuario(request.user)

    if not loja:
        return render(request, "erro_loja.html")

    if not usuario_eh_gerente_ou_admin(request.user):
        return render(request, "operacao_bloqueada.html", {
            "mensagem": "Apenas gerente ou admin podem gerenciar vendas antecipadas."
        })

    clientes = Cliente.objects.filter(loja=loja, ativo=True).order_by("nome")
    produtos = Produto.objects.filter(loja=loja).order_by("nome")
    erro = ""

    if request.method == "POST":
        acao = request.POST.get("acao", "nova")

        if acao == "utilizar":
            va_id = request.POST.get("venda_antecipada_id")
            try:
                va = VendaAntecipada.objects.get(id=va_id, loja=loja, status="pendente")
                va.status = "utilizada"
                va.utilizada_em = timezone.now()
                va.save()

                registrar_auditoria(
                    loja=loja,
                    usuario=request.user,
                    acao="Venda antecipada utilizada",
                    descricao=f"Cliente: {va.cliente.nome} | Produto: {va.produto.nome} x{va.quantidade}"
                )
            except VendaAntecipada.DoesNotExist:
                erro = "Venda antecipada não encontrada."

            return redirect("/vendas-antecipadas/")

        elif acao == "cancelar":
            va_id = request.POST.get("venda_antecipada_id")
            try:
                va = VendaAntecipada.objects.get(id=va_id, loja=loja, status="pendente")
                va.status = "cancelada"
                va.save()

                registrar_auditoria(
                    loja=loja,
                    usuario=request.user,
                    acao="Venda antecipada cancelada",
                    descricao=f"Cliente: {va.cliente.nome} | Produto: {va.produto.nome} x{va.quantidade}"
                )
            except VendaAntecipada.DoesNotExist:
                erro = "Venda antecipada não encontrada."

            return redirect("/vendas-antecipadas/")

        else:
            cliente_id = request.POST.get("cliente")
            produto_id = request.POST.get("produto")
            quantidade = request.POST.get("quantidade", "1").strip()
            valor_pago = request.POST.get("valor_pago", "").strip()
            forma_pagamento = request.POST.get("forma_pagamento", "dinheiro")
            observacoes = request.POST.get("observacoes", "").strip()

            if not cliente_id or not produto_id or not valor_pago:
                erro = "Preencha todos os campos obrigatórios."
            else:
                try:
                    cliente = Cliente.objects.get(id=cliente_id, loja=loja)
                    produto = Produto.objects.get(id=produto_id, loja=loja)

                    VendaAntecipada.objects.create(
                        loja=loja,
                        cliente=cliente,
                        produto=produto,
                        quantidade=int(quantidade),
                        valor_pago=Decimal(valor_pago.replace(",", ".")),
                        forma_pagamento=forma_pagamento,
                        observacoes=observacoes,
                        registrado_por=request.user,
                    )

                    registrar_auditoria(
                        loja=loja,
                        usuario=request.user,
                        acao="Venda antecipada registrada",
                        descricao=f"Cliente: {cliente.nome} | Produto: {produto.nome} x{quantidade} | R$ {valor_pago}"
                    )
                    return redirect("/vendas-antecipadas/")
                except (Cliente.DoesNotExist, Produto.DoesNotExist, ValueError, InvalidOperation):
                    erro = "Dados inválidos."

    status_filtro = request.GET.get("status", "pendente").strip()
    vendas_qs = VendaAntecipada.objects.filter(loja=loja).select_related("cliente", "produto")

    if status_filtro:
        vendas_qs = vendas_qs.filter(status=status_filtro)

    vendas_qs = vendas_qs.order_by("-data")

    total_pendentes = VendaAntecipada.objects.filter(loja=loja, status="pendente").count()
    valor_pendente = VendaAntecipada.objects.filter(
        loja=loja, status="pendente"
    ).aggregate(total=models.Sum("valor_pago"))["total"] or Decimal("0.00")

    return render(request, "vendas_antecipadas.html", {
        "loja": loja,
        "clientes": clientes,
        "produtos": produtos,
        "vendas": vendas_qs,
        "status_filtro": status_filtro,
        "total_pendentes": total_pendentes,
        "valor_pendente": valor_pendente,
        "erro": erro,
    })


# ========== ROTAS DE VEÍCULOS ==========

@login_required
def rotas(request):
    loja = obter_loja_usuario(request.user)
    if not loja:
        return render(request, "erro_loja.html")
    if not usuario_eh_gerente_ou_admin(request.user):
        return redirect("/")

    from datetime import date, timedelta

    filtro_data = request.GET.get("data", "")
    filtro_status = request.GET.get("status", "")

    rotas_list = VeiculoRota.objects.filter(loja=loja)

    if filtro_data:
        rotas_list = rotas_list.filter(data_rota=filtro_data)
    if filtro_status:
        rotas_list = rotas_list.filter(status=filtro_status)

    veiculos = Veiculo.objects.filter(loja=loja, status="ativo")
    entregadores = Entregador.objects.filter(loja=loja, ativo=True)

    if request.method == "POST":
        acao = request.POST.get("acao")

        if acao == "nova_rota":
            veiculo_id = request.POST.get("veiculo")
            entregador_id = request.POST.get("entregador")
            data_rota = request.POST.get("data_rota")
            nome_rota = request.POST.get("nome_rota", "")
            km_inicial = request.POST.get("km_inicial")

            try:
                veiculo = Veiculo.objects.get(id=veiculo_id, loja=loja)
                entregador = Entregador.objects.get(id=entregador_id, loja=loja) if entregador_id else None

                rota = VeiculoRota.objects.create(
                    loja=loja,
                    veiculo=veiculo,
                    entregador=entregador,
                    data_rota=data_rota,
                    nome_rota=nome_rota,
                    km_inicial=int(km_inicial) if km_inicial else None,
                )
                registrar_auditoria(request, "Nova rota criada", f"Rota {rota.data_rota} - {veiculo.placa}")
                messages.success(request, "Rota criada com sucesso!")
            except Exception as e:
                messages.error(request, f"Erro ao criar rota: {e}")

            return redirect("/rotas/")

        elif acao == "iniciar_rota":
            rota_id = request.POST.get("rota_id")
            try:
                rota = VeiculoRota.objects.get(id=rota_id, loja=loja)
                rota.status = "em_andamento"
                rota.tempo_inicio = timezone.now().time()
                rota.save()
                messages.success(request, "Rota iniciada!")
            except Exception as e:
                messages.error(request, f"Erro: {e}")
            return redirect("/rotas/")

        elif acao == "concluir_rota":
            rota_id = request.POST.get("rota_id")
            km_final = request.POST.get("km_final")
            try:
                rota = VeiculoRota.objects.get(id=rota_id, loja=loja)
                rota.status = "concluida"
                rota.tempo_fim = timezone.now().time()
                if km_final:
                    rota.km_final = int(km_final)
                    if rota.veiculo:
                        rota.veiculo.km_atual = int(km_final)
                        rota.veiculo.save()
                entregas = rota.entregas.all()
                rota.quantidade_entregas = entregas.filter(status="entregue").count()
                rota.quantidade_falhas = entregas.filter(status="falha").count()
                rota.save()
                registrar_auditoria(request, "Rota concluída", f"Rota {rota.data_rota} - {rota.veiculo.placa}")
                messages.success(request, "Rota concluída com sucesso!")
            except Exception as e:
                messages.error(request, f"Erro: {e}")
            return redirect("/rotas/")

        elif acao == "cancelar_rota":
            rota_id = request.POST.get("rota_id")
            try:
                rota = VeiculoRota.objects.get(id=rota_id, loja=loja)
                rota.status = "cancelada"
                rota.save()
                messages.success(request, "Rota cancelada.")
            except Exception as e:
                messages.error(request, f"Erro: {e}")
            return redirect("/rotas/")

    context = {
        "rotas": rotas_list[:50],
        "veiculos": veiculos,
        "entregadores": entregadores,
        "filtro_data": filtro_data,
        "filtro_status": filtro_status,
    }
    return render(request, "rotas.html", context)


@login_required
def detalhe_rota(request, rota_id):
    loja = obter_loja_usuario(request.user)
    if not loja:
        return render(request, "erro_loja.html")

    rota = get_object_or_404(VeiculoRota, id=rota_id, loja=loja)
    entregas = rota.entregas.all()
    clientes = Cliente.objects.filter(loja=loja, ativo=True)
    pedidos_disponiveis = Pedido.objects.filter(loja=loja, status__in=["novo", "preparando"])

    if request.method == "POST":
        acao = request.POST.get("acao")

        if acao == "adicionar_entrega":
            cliente_id = request.POST.get("cliente")
            pedido_id = request.POST.get("pedido")
            try:
                cliente = Cliente.objects.get(id=cliente_id, loja=loja) if cliente_id else None
                pedido = Pedido.objects.get(id=pedido_id, loja=loja) if pedido_id else None
                ordem = entregas.count() + 1
                RotaEntrega.objects.create(
                    rota=rota,
                    pedido=pedido,
                    cliente=cliente or (pedido.cliente if pedido else None),
                    ordem=ordem,
                )
                messages.success(request, "Entrega adicionada à rota!")
            except Exception as e:
                messages.error(request, f"Erro: {e}")
            return redirect(f"/rotas/{rota_id}/")

        elif acao == "marcar_entregue":
            entrega_id = request.POST.get("entrega_id")
            try:
                entrega = RotaEntrega.objects.get(id=entrega_id, rota=rota)
                entrega.status = "entregue"
                entrega.hora_entrega = timezone.now().time()
                entrega.save()
                if entrega.pedido:
                    entrega.pedido.status = "entregue"
                    entrega.pedido.data_entrega = timezone.now()
                    entrega.pedido.save()
                messages.success(request, "Entrega marcada como realizada!")
            except Exception as e:
                messages.error(request, f"Erro: {e}")
            return redirect(f"/rotas/{rota_id}/")

        elif acao == "marcar_falha":
            entrega_id = request.POST.get("entrega_id")
            obs = request.POST.get("obs_falha", "")
            try:
                entrega = RotaEntrega.objects.get(id=entrega_id, rota=rota)
                entrega.status = "falha"
                entrega.observacoes = obs
                entrega.save()
                messages.success(request, "Entrega marcada como falha.")
            except Exception as e:
                messages.error(request, f"Erro: {e}")
            return redirect(f"/rotas/{rota_id}/")

    context = {
        "rota": rota,
        "entregas": entregas,
        "clientes": clientes,
        "pedidos_disponiveis": pedidos_disponiveis,
    }
    return render(request, "detalhe_rota.html", context)


# ========== CHECKLIST DE VEÍCULOS ==========

@login_required
def checklist_veiculo(request):
    loja = obter_loja_usuario(request.user)
    if not loja:
        return render(request, "erro_loja.html")
    if not usuario_eh_gerente_ou_admin(request.user):
        return redirect("/")

    checklists = ChecklistVeiculo.objects.filter(loja=loja)[:30]
    veiculos = Veiculo.objects.filter(loja=loja, status="ativo")
    entregadores = Entregador.objects.filter(loja=loja, ativo=True)

    if request.method == "POST":
        veiculo_id = request.POST.get("veiculo")
        entregador_id = request.POST.get("entregador")
        km = request.POST.get("km_atual")

        try:
            veiculo = Veiculo.objects.get(id=veiculo_id, loja=loja)
            entregador = Entregador.objects.get(id=entregador_id, loja=loja) if entregador_id else None

            checklist = ChecklistVeiculo.objects.create(
                loja=loja,
                veiculo=veiculo,
                entregador=entregador,
                km_atual=int(km) if km else None,
                pneus_ok=request.POST.get("pneus_ok") == "on",
                freios_ok=request.POST.get("freios_ok") == "on",
                oleo_ok=request.POST.get("oleo_ok") == "on",
                agua_ok=request.POST.get("agua_ok") == "on",
                luzes_ok=request.POST.get("luzes_ok") == "on",
                limpeza_ok=request.POST.get("limpeza_ok") == "on",
                documentos_ok=request.POST.get("documentos_ok") == "on",
                extintor_ok=request.POST.get("extintor_ok") == "on",
                carga_ok=request.POST.get("carga_ok") == "on",
                observacoes=request.POST.get("observacoes", ""),
                registrado_por=request.user,
            )

            if km:
                veiculo.km_atual = int(km)
                veiculo.save()

            registrar_auditoria(request, "Checklist realizado", f"{veiculo.placa} - {checklist.percentual_ok}% OK")
            messages.success(request, f"Checklist registrado! {checklist.total_itens_ok}/{checklist.total_itens} itens OK.")
        except Exception as e:
            messages.error(request, f"Erro: {e}")

        return redirect("/checklist-veiculo/")

    context = {
        "checklists": checklists,
        "veiculos": veiculos,
        "entregadores": entregadores,
    }
    return render(request, "checklist_veiculo.html", context)


# ========== ALERTAS DE MANUTENÇÃO ==========

@login_required
def alertas_manutencao(request):
    loja = obter_loja_usuario(request.user)
    if not loja:
        return render(request, "erro_loja.html")
    if not usuario_eh_gerente_ou_admin(request.user):
        return redirect("/")

    alertas = AlertaManutencao.objects.filter(loja=loja)
    veiculos = Veiculo.objects.filter(loja=loja)

    filtro_status = request.GET.get("status", "ativo")
    if filtro_status:
        alertas = alertas.filter(status=filtro_status)

    if request.method == "POST":
        acao = request.POST.get("acao")

        if acao == "novo_alerta":
            veiculo_id = request.POST.get("veiculo")
            tipo = request.POST.get("tipo")
            titulo = request.POST.get("titulo")
            descricao = request.POST.get("descricao", "")
            km_alerta = request.POST.get("km_alerta")
            data_alerta = request.POST.get("data_alerta")

            try:
                veiculo = Veiculo.objects.get(id=veiculo_id, loja=loja)
                AlertaManutencao.objects.create(
                    loja=loja,
                    veiculo=veiculo,
                    tipo=tipo,
                    titulo=titulo,
                    descricao=descricao,
                    km_alerta=int(km_alerta) if km_alerta else None,
                    data_alerta=data_alerta if data_alerta else None,
                )
                messages.success(request, "Alerta criado com sucesso!")
            except Exception as e:
                messages.error(request, f"Erro: {e}")
            return redirect("/alertas-manutencao/")

        elif acao == "resolver_alerta":
            alerta_id = request.POST.get("alerta_id")
            try:
                alerta = AlertaManutencao.objects.get(id=alerta_id, loja=loja)
                alerta.status = "resolvido"
                alerta.resolvido_em = timezone.now()
                alerta.resolvido_por = request.user
                alerta.save()
                messages.success(request, "Alerta resolvido!")
            except Exception as e:
                messages.error(request, f"Erro: {e}")
            return redirect("/alertas-manutencao/")

        elif acao == "ignorar_alerta":
            alerta_id = request.POST.get("alerta_id")
            try:
                alerta = AlertaManutencao.objects.get(id=alerta_id, loja=loja)
                alerta.status = "ignorado"
                alerta.save()
                messages.success(request, "Alerta ignorado.")
            except Exception as e:
                messages.error(request, f"Erro: {e}")
            return redirect("/alertas-manutencao/")

    # Auto-generate alerts for upcoming maintenance
    hoje = timezone.now().date()
    for veiculo in veiculos:
        # Check document expirations
        for campo, label in [("seguro_vencimento", "Seguro"), ("ipva_vencimento", "IPVA"), ("licenciamento_vencimento", "Licenciamento")]:
            data_venc = getattr(veiculo, campo, None)
            if data_venc and data_venc <= hoje + timedelta(days=30):
                exists = AlertaManutencao.objects.filter(
                    veiculo=veiculo, tipo="documento", titulo__icontains=label, status="ativo"
                ).exists()
                if not exists:
                    AlertaManutencao.objects.create(
                        loja=loja,
                        veiculo=veiculo,
                        tipo="documento",
                        titulo=f"{label} vencendo - {veiculo.placa}",
                        descricao=f"{label} vence em {data_venc.strftime('%d/%m/%Y')}",
                        data_alerta=data_venc,
                    )

        # Check maintenance by km
        manutencoes = veiculo.manutencoes.filter(proxima_manutencao_km__isnull=False).order_by("-data")[:1]
        for m in manutencoes:
            if m.proxima_manutencao_km and veiculo.km_atual >= m.proxima_manutencao_km - 500:
                exists = AlertaManutencao.objects.filter(
                    veiculo=veiculo, tipo="km", status="ativo"
                ).exists()
                if not exists:
                    AlertaManutencao.objects.create(
                        loja=loja,
                        veiculo=veiculo,
                        tipo="km",
                        titulo=f"Manutenção por KM - {veiculo.placa}",
                        descricao=f"Próxima manutenção em {m.proxima_manutencao_km} km. KM atual: {veiculo.km_atual}",
                        km_alerta=m.proxima_manutencao_km,
                    )

    alertas_ativos = AlertaManutencao.objects.filter(loja=loja, status="ativo").count()

    context = {
        "alertas": alertas[:50],
        "veiculos": veiculos,
        "filtro_status": filtro_status,
        "alertas_ativos": alertas_ativos,
    }
    return render(request, "alertas_manutencao.html", context)


# ========== FORNECEDORES ==========

@login_required
def fornecedores(request):
    loja = obter_loja_usuario(request.user)
    if not loja:
        return render(request, "erro_loja.html")
    if not usuario_eh_admin(request.user):
        return redirect("/")

    lista = Fornecedor.objects.filter(loja=loja)
    busca = request.GET.get("q", "")
    if busca:
        lista = lista.filter(
            models.Q(nome__icontains=busca) | models.Q(cnpj__icontains=busca) | models.Q(cidade__icontains=busca)
        )

    if request.method == "POST":
        acao = request.POST.get("acao")

        if acao == "novo":
            try:
                Fornecedor.objects.create(
                    loja=loja,
                    nome=request.POST.get("nome"),
                    telefone=request.POST.get("telefone", ""),
                    email=request.POST.get("email", ""),
                    cnpj=request.POST.get("cnpj", ""),
                    endereco=request.POST.get("endereco", ""),
                    cidade=request.POST.get("cidade", ""),
                    contato=request.POST.get("contato", ""),
                    observacoes=request.POST.get("observacoes", ""),
                )
                registrar_auditoria(request, "Fornecedor criado", request.POST.get("nome"))
                messages.success(request, "Fornecedor cadastrado!")
            except Exception as e:
                messages.error(request, f"Erro: {e}")
            return redirect("/fornecedores/")

        elif acao == "inativar":
            forn_id = request.POST.get("fornecedor_id")
            try:
                forn = Fornecedor.objects.get(id=forn_id, loja=loja)
                forn.ativo = not forn.ativo
                forn.save()
                status = "ativado" if forn.ativo else "inativado"
                messages.success(request, f"Fornecedor {status}!")
            except Exception as e:
                messages.error(request, f"Erro: {e}")
            return redirect("/fornecedores/")

    context = {
        "fornecedores": lista,
        "busca": busca,
    }
    return render(request, "fornecedores.html", context)


@login_required
def editar_fornecedor(request, fornecedor_id):
    loja = obter_loja_usuario(request.user)
    if not loja:
        return render(request, "erro_loja.html")
    if not usuario_eh_admin(request.user):
        return redirect("/")

    fornecedor = get_object_or_404(Fornecedor, id=fornecedor_id, loja=loja)

    if request.method == "POST":
        fornecedor.nome = request.POST.get("nome", fornecedor.nome)
        fornecedor.telefone = request.POST.get("telefone", fornecedor.telefone)
        fornecedor.email = request.POST.get("email", fornecedor.email)
        fornecedor.cnpj = request.POST.get("cnpj", fornecedor.cnpj)
        fornecedor.endereco = request.POST.get("endereco", fornecedor.endereco)
        fornecedor.cidade = request.POST.get("cidade", fornecedor.cidade)
        fornecedor.contato = request.POST.get("contato", fornecedor.contato)
        fornecedor.observacoes = request.POST.get("observacoes", fornecedor.observacoes)
        fornecedor.save()
        messages.success(request, "Fornecedor atualizado!")
        return redirect("/fornecedores/")

    context = {"fornecedor": fornecedor}
    return render(request, "editar_fornecedor.html", context)


# ========== COMISSÕES ==========

@login_required
def comissoes(request):
    loja = obter_loja_usuario(request.user)
    if not loja:
        return render(request, "erro_loja.html")
    if not usuario_eh_admin(request.user):
        return redirect("/")

    hoje = timezone.now().date()
    mes = int(request.GET.get("mes", hoje.month))
    ano = int(request.GET.get("ano", hoje.year))

    comissoes_list = Comissao.objects.filter(loja=loja, data__month=mes, data__year=ano)
    entregadores = Entregador.objects.filter(loja=loja, ativo=True)

    if request.method == "POST":
        acao = request.POST.get("acao")

        if acao == "pagar_comissao":
            comissao_id = request.POST.get("comissao_id")
            try:
                comissao = Comissao.objects.get(id=comissao_id, loja=loja)
                comissao.status = "pago"
                comissao.pago_em = timezone.now()
                comissao.save()
                messages.success(request, "Comissão paga!")
            except Exception as e:
                messages.error(request, f"Erro: {e}")
            return redirect(f"/comissoes/?mes={mes}&ano={ano}")

        elif acao == "pagar_todas":
            entregador_id = request.POST.get("entregador_id")
            try:
                pendentes = comissoes_list.filter(entregador_id=entregador_id, status="pendente")
                pendentes.update(status="pago", pago_em=timezone.now())
                messages.success(request, "Todas as comissões foram pagas!")
            except Exception as e:
                messages.error(request, f"Erro: {e}")
            return redirect(f"/comissoes/?mes={mes}&ano={ano}")

    # Resumo por entregador
    resumo = []
    for ent in entregadores:
        comissoes_ent = comissoes_list.filter(entregador=ent)
        total = sum(c.valor_comissao for c in comissoes_ent)
        pendente = sum(c.valor_comissao for c in comissoes_ent.filter(status="pendente"))
        pago = sum(c.valor_comissao for c in comissoes_ent.filter(status="pago"))
        if total > 0:
            resumo.append({
                "entregador": ent,
                "total": total,
                "pendente": pendente,
                "pago": pago,
                "qtd": comissoes_ent.count(),
            })

    total_geral = sum(r["total"] for r in resumo)
    total_pendente = sum(r["pendente"] for r in resumo)
    total_pago = sum(r["pago"] for r in resumo)

    context = {
        "comissoes": comissoes_list[:100],
        "resumo": resumo,
        "mes": mes,
        "ano": ano,
        "total_geral": total_geral,
        "total_pendente": total_pendente,
        "total_pago": total_pago,
    }
    return render(request, "comissoes.html", context)


# ========== METAS DE VENDAS ==========

@login_required
def metas(request):
    loja = obter_loja_usuario(request.user)
    if not loja:
        return render(request, "erro_loja.html")
    if not usuario_eh_admin(request.user):
        return redirect("/")

    hoje = timezone.now().date()
    metas_list = Meta_Vendas.objects.filter(loja=loja)

    if request.method == "POST":
        mes = int(request.POST.get("mes", hoje.month))
        ano = int(request.POST.get("ano", hoje.year))
        meta_faturamento = request.POST.get("meta_faturamento", 0)
        meta_quantidade = request.POST.get("meta_quantidade_vendas", 0)
        meta_clientes = request.POST.get("meta_novos_clientes", 0)

        try:
            meta, created = Meta_Vendas.objects.update_or_create(
                loja=loja, mes=mes, ano=ano,
                defaults={
                    "meta_faturamento": Decimal(str(meta_faturamento)),
                    "meta_quantidade_vendas": int(meta_quantidade),
                    "meta_novos_clientes": int(meta_clientes),
                    "criado_por": request.user,
                }
            )
            action = "criada" if created else "atualizada"
            messages.success(request, f"Meta {action} com sucesso!")
        except Exception as e:
            messages.error(request, f"Erro: {e}")
        return redirect("/metas/")

    # Calculate progress for current month
    meta_atual = metas_list.filter(mes=hoje.month, ano=hoje.year).first()
    progresso = None
    if meta_atual:
        inicio_mes = hoje.replace(day=1)
        vendas_mes = Venda.objects.filter(loja=loja, status="ativa", data_venda__date__gte=inicio_mes)
        pedidos_mes = Pedido.objects.filter(loja=loja, status="entregue", data_pedido__date__gte=inicio_mes)

        faturamento_vendas = sum(v.total_venda() for v in vendas_mes)
        faturamento_pedidos = sum(p.total_pedido() for p in pedidos_mes)
        faturamento_total = faturamento_vendas + faturamento_pedidos

        qtd_vendas = vendas_mes.count() + pedidos_mes.count()
        novos_clientes = Cliente.objects.filter(loja=loja, criado_em__date__gte=inicio_mes).count()

        progresso = {
            "faturamento_atual": faturamento_total,
            "faturamento_meta": meta_atual.meta_faturamento,
            "faturamento_pct": round((float(faturamento_total) / float(meta_atual.meta_faturamento)) * 100, 1) if meta_atual.meta_faturamento > 0 else 0,
            "vendas_atual": qtd_vendas,
            "vendas_meta": meta_atual.meta_quantidade_vendas,
            "vendas_pct": round((qtd_vendas / meta_atual.meta_quantidade_vendas) * 100, 1) if meta_atual.meta_quantidade_vendas > 0 else 0,
            "clientes_atual": novos_clientes,
            "clientes_meta": meta_atual.meta_novos_clientes,
            "clientes_pct": round((novos_clientes / meta_atual.meta_novos_clientes) * 100, 1) if meta_atual.meta_novos_clientes > 0 else 0,
        }

    context = {
        "metas": metas_list[:12],
        "meta_atual": meta_atual,
        "progresso": progresso,
        "hoje": hoje,
    }
    return render(request, "metas.html", context)


# ========== MÉTRICAS DE ENTREGADORES ==========

@login_required
def metricas_entregadores(request):
    loja = obter_loja_usuario(request.user)
    if not loja:
        return render(request, "erro_loja.html")
    if not usuario_eh_admin(request.user):
        return redirect("/")

    hoje = timezone.now().date()
    periodo = request.GET.get("periodo", "mes")

    if periodo == "hoje":
        data_inicio = hoje
    elif periodo == "semana":
        data_inicio = hoje - timedelta(days=7)
    elif periodo == "mes":
        data_inicio = hoje.replace(day=1)
    else:
        data_inicio = hoje.replace(day=1)

    entregadores = Entregador.objects.filter(loja=loja, ativo=True)
    metricas = []

    for ent in entregadores:
        pedidos_ent = Pedido.objects.filter(
            loja=loja,
            entregador=ent,
            data_pedido__date__gte=data_inicio,
        )
        entregues = pedidos_ent.filter(status="entregue")
        cancelados = pedidos_ent.filter(status="cancelado")

        total = pedidos_ent.count()
        total_entregues = entregues.count()
        total_cancelados = cancelados.count()
        taxa = round((total_entregues / total) * 100, 1) if total > 0 else 100

        valor_total = sum(p.total_pedido() for p in entregues)

        rotas_ent = VeiculoRota.objects.filter(
            loja=loja, entregador=ent, data_rota__gte=data_inicio
        )
        km_total = sum(r.km_rodado for r in rotas_ent)
        total_rotas = rotas_ent.count()

        comissao_total = Decimal("0")
        if ent.comissao_percentual and ent.comissao_percentual > 0:
            comissao_total = valor_total * ent.comissao_percentual / 100

        metricas.append({
            "entregador": ent,
            "total_pedidos": total,
            "total_entregues": total_entregues,
            "total_cancelados": total_cancelados,
            "taxa_sucesso": taxa,
            "valor_total": valor_total,
            "km_total": km_total,
            "total_rotas": total_rotas,
            "comissao": comissao_total,
        })

    metricas.sort(key=lambda x: x["total_entregues"], reverse=True)

    context = {
        "metricas": metricas,
        "periodo": periodo,
        "data_inicio": data_inicio,
    }
    return render(request, "metricas_entregadores.html", context)


# ========== NOTIFICAÇÕES ==========

@login_required
def notificacoes(request):
    notifs_qs = NotificacaoSistema.objects.filter(usuario=request.user)

    if request.method == "POST":
        acao = request.POST.get("acao")
        if acao == "marcar_lida":
            notif_id = request.POST.get("notif_id")
            try:
                notif = NotificacaoSistema.objects.get(id=notif_id, usuario=request.user)
                notif.lida = True
                notif.save()
            except:
                pass
            return redirect("/notificacoes/")
        elif acao == "marcar_todas":
            notifs_qs.filter(lida=False).update(lida=True)
            messages.success(request, "Todas as notificações marcadas como lidas.")
            return redirect("/notificacoes/")

    nao_lidas = notifs_qs.filter(lida=False).count()

    context = {
        "notificacoes": notifs_qs[:50],
        "nao_lidas": nao_lidas,
    }
    return render(request, "notificacoes.html", context)


@login_required
def notificacoes_json(request):
    nao_lidas = NotificacaoSistema.objects.filter(usuario=request.user, lida=False).count()
    return JsonResponse({"nao_lidas": nao_lidas})


@login_required
def trocar_loja(request):
    """Permite ao usuário com múltiplos perfis alternar entre lojas."""
    if request.method == "POST":
        loja_id = request.POST.get("loja_id")
        if loja_id:
            tem_acesso = PerfilUsuario.objects.filter(
                user=request.user, loja_id=loja_id
            ).exists()
            if tem_acesso or request.user.is_staff:
                request.session['loja_ativa_id'] = int(loja_id)
    return redirect(request.POST.get("next", "/"))
